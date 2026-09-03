"""Clean the Australian higher education finance sources into partitioned parquet.

Two sources, six output tables:

  income_statement   |
  balance_sheet      | Finance Publication, one row per provider-year-line item
  equity_movement    |
  cash_flow          |
  research_income      HERDC, one row per provider-year-category-sub-category
  line_item            which line items appear in which statement and years

The Finance Publication ships wide pivot exports of the department's HAROLD
cube: providers across the columns, grouped under a sparse state header, with
statement line items down the rows. Up to 2021 each state group ends with a
subtotal column and the sheet ends with a national total; those aggregates are
dropped here because summing the providers reproduces them.

Provider identity comes from HERDC's HEP code via `providers.ProviderIndex` --
the finance series renames providers over its span and carries no code of its
own. Every money column is written in Australian dollars: HERDC already reports
dollars, and the finance statements are scaled up from the thousands the source
prints, so the two can be compared without a hidden factor of a thousand.
"""

from __future__ import annotations

import collections
import json
import os
import pathlib
import re
import shutil
import sys

import openpyxl  # pyrefly: ignore [untyped-import]
import pyarrow as pa
import pyarrow.parquet as pq

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from providers import (  # pyrefly: ignore [missing-import]
    STATE_GROUPS,
    ProviderIndex,
)

DATA_DIR = pathlib.Path(
    os.environ.get(
        "AU_DOE_HEF_DATA",
        pathlib.Path.home()
        / "Downloads/au_doe_higher_education_finances_data",
    )
)
INPUT_DIR = DATA_DIR / "input"
FINANCE_DIR = INPUT_DIR / "finance"
OUTPUT_DIR = DATA_DIR / "output"

# Sheet name -> output table. The department retitles these sheets between
# releases ("Financial Perf - Total", "Financial Performance - Total",
# "Financial Performance"; "Comp Income & Equity - Total" versus
# "Comprehensive Income & Equity"), so the name is normalised to lowercase
# alphanumerics before matching. The needles must therefore be written in that
# same normalised form -- "income & equity" would never match, because the
# ampersand is gone by the time the comparison happens.
STATEMENT_SHEETS = [
    ("dual sector", "dual"),  # must be tested before the plain "perf" rule
    ("financial perf", "income_statement"),
    ("financial pos", "balance_sheet"),
    ("income equity", "equity_movement"),
    ("income and equity", "equity_movement"),
    ("cash flow", "cash_flow"),
]

# The source writes the sector split three ways across sheets and releases.
INSTITUTION_TYPE = {
    "hed": "higher_education",
    "vet": "vocational_education",
    "tafe": "vocational_education",
    "total institution": "total",
}

# The Finance Publication prints every statement in thousands of dollars (each
# sheet title ends "($'000)"), while HERDC reports research income in dollars.
# Leaving that split in place puts two money columns on two scales inside one
# dataset, so anything comparing them -- research income as a share of revenue,
# say -- is wrong by a factor of a thousand with nothing to signal it. The
# statements are therefore scaled to dollars here. The conversion is exact:
# every published figure is a whole number of thousands.
THOUSANDS_TO_DOLLARS = 1000


def text(value) -> str:
    return "" if value is None else str(value).strip()


def classify_sheet(name: str) -> str | None:
    """Map a sheet name to an output table.

    Separators are normalised because the department alternates between
    "Dual Sectors", "Dual-Sectors" and "Dual Sector" across releases; matching
    the raw string sends the 2018 and 2020 dual-sector sheets into the income
    statement branch, where they parse to nothing.
    """
    lowered = re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()
    for needle, table in STATEMENT_SHEETS:
        if needle in lowered:
            return table
    return None


def to_number(raw: str) -> float | None:
    """Parse a cell to a number, treating the source's blanks and dashes as null."""
    cleaned = raw.replace(",", "").replace("$", "").strip()
    if cleaned in ("", "-", "--", "n/a", "na", "np", ".."):
        return None
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return None


def sheet_rows(sheet) -> list[list[str]]:
    return [
        [text(c) for c in row] for row in sheet.iter_rows(values_only=True)
    ]


def workbook_sheets(path: pathlib.Path) -> dict[str, list[list[str]]]:
    """Every sheet of a workbook as rows of strings, in either Excel format.

    Releases up to 2016 ship the legacy .xls format, which openpyxl cannot open
    at all; xlrd reads it. The layout is identical across the two formats, so
    only the reader differs. xlrd surfaces every number as a float, so a cell
    arrives as "550016.0" -- to_number parses that fine, and write_partitioned
    puts integral values back to int before serialising.

    The format is sniffed from the file's magic bytes rather than its
    extension, because the department serves legacy workbooks from URLs and
    filenames that claim otherwise.
    """
    if not path.read_bytes()[:2].startswith(b"PK"):
        import xlrd  # pyrefly: ignore [untyped-import]

        book = xlrd.open_workbook(str(path))
        return {
            sheet.name: [
                [text(cell.value) for cell in sheet.row(i)]
                for i in range(sheet.nrows)
            ]
            for sheet in (book.sheet_by_index(i) for i in range(book.nsheets))
        }

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return {name: sheet_rows(book[name]) for name in book.sheetnames}


def find_header(rows: list[list[str]]) -> tuple[int, int] | None:
    """Locate (state row, first data column) in a standard wide statement sheet."""
    for i, row in enumerate(rows[:20]):
        hits = [j for j, v in enumerate(row) if v in STATE_GROUPS]
        if len(hits) >= 2:
            return i, min(hits)
    return None


def labels_for(row: list[str], first_col: int) -> tuple[str, str]:
    """Published label and internal cube name for a data row.

    The label columns are everything before the first provider column. Where
    there are two, the first is the HAROLD cube's internal member name ("DEEWR
    Research Grants", "State Govt Total") and the second is the published label
    ("Education Research Grants", "State and Local Government Financial
    Assistance"). The 2023 release ships a single label column carrying the
    published label and no cube name at all.

    So the published label is always the LAST label column, never the first.
    Reading column 0 as the label silently mixes two different vocabularies, and
    reading column 1 as a second label turns 2023's first provider value into a
    label.
    """
    labels = [row[i] for i in range(min(first_col, len(row))) if row[i]]
    if not labels:
        return "", ""
    if len(labels) == 1:
        return labels[0], ""
    return labels[-1], labels[0]


def parse_statement_sheet(rows, year, table, index, report) -> list[dict]:
    """One wide statement sheet -> long records, aggregates dropped."""
    header = find_header(rows)
    if header is None:
        report["skipped"].append(f"{year}/{table}: no state header row")
        return []
    state_row, first_col = header
    states, names = rows[state_row], rows[state_row + 1]

    # The state header is a sparse group label; carry it across its columns so a
    # column whose provider label equals its own group is recognised as that
    # group's subtotal.
    group, columns = "", {}
    for col in range(first_col, max(len(states), len(names))):
        if col < len(states) and states[col]:
            group = states[col]
        label = names[col] if col < len(names) else ""
        if not label:
            continue
        if index.is_aggregate(label) or label == group:
            continue
        code = index.code(label)
        if code is not None:
            columns[col] = code

    records = []
    line_number = 0
    for row in rows[state_row + 2 :]:
        if len(row) <= first_col:
            continue
        line_item, internal = labels_for(row, first_col)
        if not line_item:
            continue
        # The source repeats labels within a statement -- "Other Comprehensive
        # Income" heads two separate sections of the equity statement, and the
        # cube name repeats with it -- so no combination of labels identifies a
        # row. The position in the statement does, and it carries the
        # presentation order a reader needs to render the statement.
        line_number += 1
        for col, code in columns.items():
            if col >= len(row):
                continue
            value = to_number(row[col])
            if value is None:
                continue
            records.append(
                {
                    "year": year,
                    "hep_code": code,
                    "institution_type": "total",
                    "line_number": line_number,
                    "line_item": line_item,
                    "line_item_internal": internal,
                    "value": value * THOUSANDS_TO_DOLLARS,
                }
            )
    return records


def parse_dual_sector_sheet(rows, year, index, report) -> list[dict]:
    """The dual-sector sheet splits income statements into HED / VET / total.

    Its layout differs from the other sheets: providers sit on one row as a
    sparse group header over three sector columns each, with the sector labels
    on the row below and no state grouping at all.
    """
    sector_row = None
    for i, row in enumerate(rows[:20]):
        if sum(v.lower() in INSTITUTION_TYPE for v in row) >= 2:
            sector_row = i
            break
    if sector_row is None or sector_row == 0:
        report["skipped"].append(f"{year}/dual_sector: no sector header row")
        return []

    provider_row = rows[sector_row - 1]
    sectors = rows[sector_row]
    first_col = min(
        j for j, v in enumerate(sectors) if v.lower() in INSTITUTION_TYPE
    )

    provider, columns = "", {}
    for col in range(first_col, len(sectors)):
        if col < len(provider_row) and provider_row[col]:
            provider = provider_row[col]
        sector = INSTITUTION_TYPE.get(sectors[col].lower())
        if not sector or not provider:
            continue
        # Each provider block here is HED / VET / Total Institution, and that
        # third column repeats what the main sheet already reports for the same
        # provider and line. Taking it would duplicate every dual-sector
        # provider's whole income statement.
        if sector == "total":
            continue
        code = index.code(provider)
        if code is not None:
            columns[col] = (code, sector)

    records = []
    line_number = 0
    for row in rows[sector_row + 1 :]:
        if len(row) <= first_col:
            continue
        line_item, internal = labels_for(row, first_col)
        if not line_item:
            continue
        line_number += 1
        for col, (code, sector) in columns.items():
            if col >= len(row):
                continue
            value = to_number(row[col])
            if value is None:
                continue
            records.append(
                {
                    "year": year,
                    "hep_code": code,
                    "institution_type": sector,
                    "line_number": line_number,
                    "line_item": line_item,
                    "line_item_internal": internal,
                    "value": value * THOUSANDS_TO_DOLLARS,
                }
            )
    return records


def read_herdc(path: pathlib.Path) -> tuple[dict, list[dict]]:
    """HERDC providers and the long research income records.

    Sheet 3 is already one row per provider-year-sub-category, so it is used
    directly; sheet 1 supplies each provider's state and cohort.
    """
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)

    providers: dict[str, dict] = {}
    for row in book["1. Summary by Category"].iter_rows(
        min_row=4, values_only=True
    ):
        code = text(row[0])
        if not code or not text(row[2]).isdigit():
            continue
        year = int(text(row[2]))
        meta = providers.setdefault(
            code,
            {
                "name": text(row[1]),
                "state": text(row[3]),
                "cohort": text(row[4]),
                "first_year": year,
                "last_year": year,
            },
        )
        meta["first_year"] = min(meta["first_year"], year)
        meta["last_year"] = max(meta["last_year"], year)
        # Later rows are more recent releases; keep the newest naming.
        if year >= meta["last_year"]:
            meta["name"], meta["state"], meta["cohort"] = (
                text(row[1]),
                text(row[3]),
                text(row[4]),
            )

    records = []
    for row in book["3. Accessible column format"].iter_rows(
        min_row=4, values_only=True
    ):
        code = text(row[0])
        if not code or not text(row[2]).isdigit():
            continue
        # A blank amount means the sub-category was not in use that year, which
        # the source documents in its own historical sub-category map. That is
        # not the same as zero, so the row is kept with a null amount rather
        # than dropped -- otherwise the panel silently loses 2,790 rows and
        # users cannot tell "not collected" from "collected as nil".
        records.append(
            {
                "year": int(text(row[2])),
                "hep_code": code,
                "category": text(row[5]),
                "sub_category": text(row[6]),
                "amount": to_number(text(row[7])),
            }
        )
    return providers, records


def write_partitioned(
    records: list[dict], table: str, schema: pa.Schema
) -> int:
    """Write one table as year-partitioned parquet, all columns STRING.

    Staging is all-STRING by house convention and the dbt model safe_casts each
    column, so the schema here fixes column order, not types. Values pass through
    their real Python types before being cast via arrow, so a year serialises as
    "2024" rather than "2024.0" and a null stays null rather than becoming "nan".
    """
    if not records:
        return 0
    by_year: dict[int, list[dict]] = collections.defaultdict(list)
    for record in records:
        by_year[record["year"]].append(record)

    fields = [f.name for f in schema]
    total = 0
    for year, rows in sorted(by_year.items()):
        columns = {}
        for name in fields:
            values = [r.get(name) for r in rows]
            if name in ("value", "amount"):
                # Integral in the source; keep them that way so the string form
                # has no spurious decimal point.
                values = [
                    None
                    if v is None
                    else (int(v) if float(v).is_integer() else v)
                    for v in values
                ]
            columns[name] = pa.array(
                [None if v is None else str(v) for v in values],
                type=pa.string(),
            )
        table_out = pa.Table.from_pydict(columns, schema=schema)
        dest = OUTPUT_DIR / table / f"year={year}"
        dest.mkdir(parents=True, exist_ok=True)
        pq.write_table(table_out, dest / "data.parquet", compression="snappy")
        total += table_out.num_rows
    return total


STATEMENT_SCHEMA = pa.schema(
    [
        ("year", pa.string()),
        ("hep_code", pa.string()),
        ("institution_type", pa.string()),
        ("line_number", pa.string()),
        ("line_item", pa.string()),
        ("line_item_internal", pa.string()),
        ("value", pa.string()),
    ]
)

RESEARCH_SCHEMA = pa.schema(
    [
        ("year", pa.string()),
        ("hep_code", pa.string()),
        ("category", pa.string()),
        ("sub_category", pa.string()),
        ("amount", pa.string()),
    ]
)

# The only genuinely coded column in the dataset. Everything else stores
# readable labels, so it needs a description, not a dictionary entry.
DICTIONARY = [
    ("research_income", "category", "1", "Australian competitive grants"),
    (
        "research_income",
        "category",
        "2",
        "Other public sector research income",
    ),
    ("research_income", "category", "3", "Industry and other research income"),
    (
        "research_income",
        "category",
        "4",
        "Cooperative Research Centre (CRC) research income",
    ),
]


def main() -> None:
    # Start from an empty output tree so a partition that disappears between
    # runs cannot survive and be uploaded alongside the current data.
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report: dict = {"skipped": [], "years": {}, "rows": {}}

    herdc_path = INPUT_DIR / "research_income_time_series.xlsx"
    providers, research = read_herdc(herdc_path)
    index = ProviderIndex(providers)
    print(
        f"HERDC: {len(providers)} providers, {len(research)} research income rows"
    )

    tables: dict[str, list[dict]] = collections.defaultdict(list)
    workbooks = sorted(
        list(FINANCE_DIR.glob("finance_*.xlsx"))
        + list(FINANCE_DIR.glob("finance_*.xls"))
    )
    for path in workbooks:
        matched = re.search(r"(\d{4})", path.name)
        if matched is None:
            report["skipped"].append(f"{path.name}: no year in filename")
            continue
        year = int(matched.group(1))
        per_year: dict[str, int] = {}
        for name, rows in workbook_sheets(path).items():
            if not any(any(cell for cell in row) for row in rows):
                continue  # releases occasionally ship an empty trailing "Sheet1"
            table = classify_sheet(name)
            if table is None:
                report["skipped"].append(
                    f"{year}: unclassified sheet {name!r}"
                )
                continue
            if table == "dual":
                records = parse_dual_sector_sheet(rows, year, index, report)
                tables["income_statement"].extend(records)
                key = "income_statement_dual"
            else:
                records = parse_statement_sheet(
                    rows, year, table, index, report
                )
                tables[table].extend(records)
                key = table
            # Accumulate: two sheets can feed one table, and assigning here
            # would let the second overwrite the first's count.
            per_year[key] = per_year.get(key, 0) + len(records)
        report["years"][year] = per_year
        print(
            f"{year}: "
            + ", ".join(f"{k}={v}" for k, v in sorted(per_year.items()))
        )

    if index.unmatched:
        print("\n!! provider labels that did not resolve to a HEP code:")
        for label in sorted(index.unmatched):
            print(f"   {label!r}")
        raise SystemExit(
            "Refusing to write output: an unresolved label silently drops a "
            "provider's entire history. Add it to providers.ALIASES."
        )

    for table, records in sorted(tables.items()):
        report["rows"][table] = write_partitioned(
            records, table, STATEMENT_SCHEMA
        )
    report["rows"]["research_income"] = write_partitioned(
        research, "research_income", RESEARCH_SCHEMA
    )

    # Which line items exist in which statement and years. The source relabels
    # items across its span, so this is the map users need to read a long series.
    seen: dict[tuple[str, str], set[int]] = collections.defaultdict(set)
    for table, records in tables.items():
        for record in records:
            seen[(table, record["line_item"])].add(record["year"])
    line_items = [
        {
            "statement": statement,
            "line_item": item,
            "first_year": min(years),
            "last_year": max(years),
            "n_years": len(years),
        }
        for (statement, item), years in sorted(seen.items())
    ]
    (OUTPUT_DIR / "line_item").mkdir(parents=True, exist_ok=True)
    pq.write_table(
        pa.Table.from_pydict(
            {
                k: pa.array([str(r[k]) for r in line_items], type=pa.string())
                for k in (
                    "statement",
                    "line_item",
                    "first_year",
                    "last_year",
                    "n_years",
                )
            },
            schema=pa.schema(
                [
                    (c, pa.string())
                    for c in (
                        "statement",
                        "line_item",
                        "first_year",
                        "last_year",
                        "n_years",
                    )
                ]
            ),
        ),
        OUTPUT_DIR / "line_item/data.parquet",
        compression="snappy",
    )
    report["rows"]["line_item"] = len(line_items)

    research_years = sorted({r["year"] for r in research})
    coverage = (
        f"{research_years[0]}(1){research_years[-1]}" if research_years else ""
    )
    dictionary_fields = (
        "id_tabela",
        "nome_coluna",
        "chave",
        "cobertura_temporal",
        "valor",
    )
    dictionary_rows = [
        {
            "id_tabela": table,
            "nome_coluna": col,
            "chave": key,
            "cobertura_temporal": coverage,
            "valor": label,
        }
        for table, col, key, label in DICTIONARY
    ]
    (OUTPUT_DIR / "dicionario").mkdir(parents=True, exist_ok=True)
    pq.write_table(
        pa.Table.from_pydict(
            {
                f: pa.array([r[f] for r in dictionary_rows], type=pa.string())
                for f in dictionary_fields
            },
            schema=pa.schema([(f, pa.string()) for f in dictionary_fields]),
        ),
        OUTPUT_DIR / "dicionario/data.parquet",
        compression="snappy",
    )
    report["rows"]["dicionario"] = len(dictionary_rows)

    providers_out = [
        {"hep_code": code, **meta} for code, meta in sorted(providers.items())
    ]
    (DATA_DIR / "providers.json").write_text(
        json.dumps(providers_out, indent=2)
    )

    # The provider directory table. It belongs to br_bd_diretorios_au rather
    # than to this dataset -- provider identity is a shared entity that any
    # future Australian higher education dataset will need -- so it is written
    # to its own tree for a separate upload. Only identity attributes go in;
    # the years a provider reported are a fact of research_income, not of the
    # provider, and stay there.
    directory_dir = DATA_DIR / "directory_output/higher_education_provider"
    directory_dir.mkdir(parents=True, exist_ok=True)
    fields = ("hep_code", "name", "abbreviation_state", "cohort")
    rows = [
        {
            "hep_code": code,
            "name": meta["name"],
            "abbreviation_state": meta["state"],
            "cohort": meta["cohort"],
        }
        for code, meta in sorted(providers.items())
    ]
    pq.write_table(
        pa.Table.from_pydict(
            {
                f: pa.array([r[f] for r in rows], type=pa.string())
                for f in fields
            },
            schema=pa.schema([(f, pa.string()) for f in fields]),
        ),
        directory_dir / "data.parquet",
        compression="snappy",
    )
    report["rows"]["higher_education_provider (directory)"] = len(rows)
    (DATA_DIR / "clean_report.json").write_text(json.dumps(report, indent=2))

    print("\nrows written:")
    for table, count in sorted(report["rows"].items()):
        print(f"  {table:20s} {count:>9,}")
    if report["skipped"]:
        print(f"\nskipped ({len(report['skipped'])}):")
        for note in report["skipped"]:
            print(f"  {note}")


if __name__ == "__main__":
    main()
