"""Probe the layout of every Finance Publication workbook.

The wide sheets are pivot exports from the department's HAROLD cube. Their
preamble height, the row carrying state names, and whether a sheet repeats its
institution block for a second (prior-year) scenario all vary across releases.
This dumps what each sheet actually looks like so the parser can be written
against observed structure rather than assumed structure.
"""

from __future__ import annotations

import os
import pathlib

import openpyxl  # pyrefly: ignore [untyped-import]

DATA_DIR = pathlib.Path(
    os.environ.get(
        "AU_DOE_HEF_DATA",
        pathlib.Path.home()
        / "Downloads/au_doe_higher_education_finances_data",
    )
)
FINANCE_DIR = DATA_DIR / "input/finance"

STATES = {
    "New South Wales",
    "Victoria",
    "Queensland",
    "South Australia",
    "Western Australia",
    "Tasmania",
    "Northern Territory",
    "Australian Capital Territory",
    "Multi-State",
    "Multi State",
}


def cell(value) -> str:
    return "" if value is None else str(value).strip()


def probe(path: pathlib.Path) -> None:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n########## {path.name}")
    for name in book.sheetnames:
        sheet = book[name]
        rows = [
            [cell(c) for c in r] for r in sheet.iter_rows(values_only=True)
        ]
        width = max((len(r) for r in rows), default=0)

        meta = {}
        state_row = institution_row = None
        for i, row in enumerate(rows[:20]):
            for j, val in enumerate(row):
                if val.rstrip(":") in (
                    "Cube",
                    "Year",
                    "Scenario",
                    "Institution Type",
                ):
                    meta[val.rstrip(":")] = [
                        (k, v)
                        for k, v in enumerate(row[j + 1 :], start=j + 1)
                        if v
                    ]
            if state_row is None and sum(v in STATES for v in row) >= 2:
                state_row = i
                institution_row = i + 1

        print(f"--- {name}  rows={len(rows)} cols={width}")
        for key in ("Year", "Scenario", "Institution Type"):
            if key in meta:
                print(f"    {key}: {meta[key]}")
        if state_row is None:
            print("    !! no state header row found")
            continue

        assert institution_row is not None
        states = [(j, v) for j, v in enumerate(rows[state_row]) if v]
        insts = [(j, v) for j, v in enumerate(rows[institution_row]) if v]
        print(f"    state row={state_row} n={len(states)} first={states[:3]}")
        print(
            f"    inst  row={institution_row} n={len(insts)} "
            f"cols[{insts[0][0] if insts else '-'}..{insts[-1][0] if insts else '-'}]"
        )

        # How many label columns precede the first institution column?
        first_col = insts[0][0] if insts else 2
        data = [r for r in rows[institution_row + 1 :] if any(r[:first_col])]
        differing = sum(
            1 for r in data if len(r) > 1 and r[0] and r[1] and r[0] != r[1]
        )
        print(
            f"    label cols=0..{first_col - 1}  data rows={len(data)}  "
            f"rows where colA != colB: {differing}"
        )
        for r in data[:3]:
            print(f"      {r[:first_col]} -> {r[first_col : first_col + 3]}")


def main() -> None:
    for path in sorted(FINANCE_DIR.glob("finance_*.xlsx")):
        probe(path)


if __name__ == "__main__":
    main()
