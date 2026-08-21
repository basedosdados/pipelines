"""Parse the PIAAC international codebooks into typed column specifications.

The two codebooks are the schema source of truth: their variable counts match the
CSV headers exactly (1,328 for Cycle 1, 2,483 for Cycle 2), so every column in the
data is described here.

Typing does NOT follow the codebook's own `Level` field. That field is inconsistent
across cycles -- AGE_R is `Ratio` in Cycle 1 but `Nominal` in Cycle 2, and Cycle 2
marks its timing variables `Not defined` -- so trusting it would silently turn a
quantity into a category or the reverse. Instead, a column is numeric only when it
matches an explicit rule that also names its measurement unit, per
`.claude/rules/data-basis-style.md`. Anything unclassified raises rather than
defaulting, so gaps surface here instead of in BigQuery.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Domains whose variables are per-item measures rather than per-respondent facts.
# The five cognitive domains are the bulk of it; the tutorial and effort-question
# blocks carry the identical suffix structure and belong in the same long table.
ITEM_DOMAINS = {
    "Tutorial",
    "Effort questions",
}

COGNITIVE_DOMAINS = {
    "Literacy (computer)",
    "Literacy (paper)",
    "Numeracy (computer)",
    "Numeracy (paper)",
    "Problem solving (computer)",
    "Adaptive problem solving (computer)",
    "Reading components (paper)",
    "Reading components (computer)",
    "Numeracy components (computer)",
}
ITEM_DOMAINS |= COGNITIVE_DOMAINS

# Suffix -> (column name, bigquery type, measurement unit). Matched longest-first so
# that VS is not read as V+S and TT is not read as T+T.
ITEM_MEASURES: list[tuple[str, str, str, str]] = [
    ("VS", "n_short_visits", "INT64", "visit"),
    ("TT", "timing_seconds", "FLOAT64", "second"),
    ("A", "n_actions", "INT64", "action"),
    ("F", "timing_first_action_seconds", "FLOAT64", "second"),
    ("R", "raw_response", "STRING", ""),
    ("S", "scored_response", "STRING", ""),
    ("T", "timing_seconds", "FLOAT64", "second"),
    ("V", "n_visits", "INT64", "visit"),
]

# Numeric columns that genuinely have no measurement unit. These are the only
# columns allowed to ship unitless, and each carries an explanation in
# `observations`. Everything else must match a UNIT_RULES pattern.
# Numeric columns whose unit cannot be read off the label. A blank unit means the
# quantity is genuinely dimensionless -- a sampling weight, a Fay factor, a
# z-score -- and every such column explains itself in `observations`. Nothing
# ships unitless by accident.
UNITLESS_NUMERIC: list[tuple[re.Pattern, str, str]] = [
    (
        re.compile(r"^SPFWT\d+$", re.I),
        "",
        "Dimensionless final or replicate sampling weight. SPFWT0 is the final "
        "weight; SPFWT1-80 are the 80 replicate weights used for variance estimation",
    ),
    (
        re.compile(r"^PV(LIT|NUM|PSL|APS)\d+$", re.I),
        "score_point",
        "Plausible value on the PIAAC proficiency scale, which runs from 0 to 500. "
        "Use all ten plausible values together, never a single one",
    ),
    (
        re.compile(
            r"_WLE_CA$|^(LEARNATWORK|READYTOLEARN|ICTHOME|ICTWORK|INFLUENCE|"
            r"NUMHOME|NUMWORK|PLANNING|READHOME|READWORK|TASKDISC|WRITHOME|"
            r"WRITWORK)$",
            re.I,
        ),
        "score_point",
        "Weighted likelihood estimate on a derived skill-use scale",
    ),
    (
        re.compile(r"^VEFAYFAC$", re.I),
        "",
        "Dimensionless Fay factor used in replicate-weight variance estimation",
    ),
    (
        re.compile(
            r"^(CBA_CORE_STAGE\d_SCORE|PPC_SCORE|PRC_(PV|SP|PC)_SCR)$", re.I
        ),
        "score_point",
        "Raw number of correct responses, not a scaled proficiency score",
    ),
]

# Same idea, but matched against the label, for families whose names carry no
# common stem. Cycle 2 added all of these; Cycle 1 has none.
UNITLESS_NUMERIC_BY_LABEL: list[tuple[re.Pattern, str, str]] = [
    (
        re.compile(r"z-score", re.I),
        "",
        "Dimensionless standardised score (mean 0, standard deviation 1) on a "
        "BFI-2 personality inventory dimension or facet",
    ),
    (
        re.compile(r"^index of\b", re.I),
        "score_point",
        "Derived index of skill use. Cycle 2 only, and not comparable with the "
        "Cycle 1 scale of the same name",
    ),
    (
        re.compile(r"standard error", re.I),
        "score_point",
        "Standard error of the derived index, on the same scale as the index",
    ),
]

# Ordered: first match wins, so specific patterns precede general ones. Matched
# against "<NAME> <label>".
UNIT_RULES: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\bppp\b|purchasing power", re.I), "usd"),
    (re.compile(r"percentile rank|\bpercent", re.I), "percent"),
    (
        re.compile(
            r"earnings|gross pay|\bincome\b|\bwage|payments amount|salary interval",
            re.I,
        ),
        "",
    ),
    (re.compile(r"months elapsed|\bmonths\b", re.I), "month"),
    (re.compile(r"\bweeks\b", re.I), "week"),
    (re.compile(r"\bdays\b", re.I), "day"),
    (re.compile(r"hours per|hours/week|\bhours\b|hour of", re.I), "hour"),
    (re.compile(r"\btimer\b|\btiming\b|seconds", re.I), "second"),
    (re.compile(r"\bage\b|age of|age at|age when", re.I), "year"),
    # `-\s*year$` anchors to the end of the label so that "... - Year" is a year but
    # "Activities - Last year - ... - Count" still falls through to the count rule.
    (
        re.compile(
            r"year of|year when|years of|\byears\b|into years|-\s*year$", re.I
        ),
        "year",
    ),
    (
        re.compile(
            r"people in household|number of people|number living in household",
            re.I,
        ),
        "person",
    ),
    (
        re.compile(r"number of children|number of siblings|^J2?_Q03b$", re.I),
        "person",
    ),
    (re.compile(r"number of actions", re.I), "action"),
    (re.compile(r"\bvisit\b", re.I), "visit"),
    (re.compile(r"\bcount\b|how many|number of", re.I), ""),
]

# PIAAC names every item measure with a trailing parenthetical. It states what the
# variable actually counts, which the surrounding item title often contradicts.
MEASURE_PARENTHETICAL = re.compile(
    r"\((Number of Actions|Timing First Action|Timing|Short Visit|Visit)\)\s*$",
    re.I,
)

# Identifiers and free text: digits, but arithmetic on them is meaningless.
STRING_BY_NAME = re.compile(
    r"^(SEQID|CNTRYID|CNTRYID_E|CNTRY|CNTRY_E|VARSTRAT|VARUNIT|VEMETHOD|VEMETHODN|"
    r"VENREPS|LANGUAGE|EARNFLAG|.*DCL)$",
    re.I,
)


@dataclass
class Variable:
    cycle: str
    name: str
    label: str
    level: str
    decimals: int
    domain: str
    value_scheme: str
    missing_scheme: str
    missing_scheme_sas: str = ""

    @property
    def is_item(self) -> bool:
        return self.domain in ITEM_DOMAINS

    @property
    def has_value_scheme(self) -> bool:
        return bool(self.value_scheme.strip())


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def load_codebook(path: Path, cycle: str) -> list[Variable]:
    """Both codebooks carry the same fields under different sheet and column layouts."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    variables = []
    if cycle == "1":
        # Sequence, Name, Label, Type, Level, Width, Decimals, Domain, Value scheme
        for row in workbook["Variables"].iter_rows(
            min_row=2, values_only=True
        ):
            if row[1] is None:
                continue
            variables.append(
                Variable(
                    "1",
                    _text(row[1]),
                    _text(row[2]),
                    _text(row[4]),
                    int(row[6] or 0),
                    _text(row[7]),
                    _text(row[8]),
                    "",
                )
            )
    else:
        # Sequence, Variable, Label, Level, Width, Decimals, RangeMin, RangeMax,
        # ValueScheme, MissingSPSS, MissingSAS, Domain, Comment
        for row in workbook["PUF"].iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            variables.append(
                Variable(
                    "2",
                    _text(row[1]),
                    _text(row[2]),
                    _text(row[3]),
                    int(row[5] or 0),
                    _text(row[11]),
                    _text(row[8]),
                    _text(row[9]),
                    _text(row[10]),
                )
            )
    workbook.close()
    return variables


def split_item(name: str) -> tuple[str, str, str, str] | None:
    """Split a cognitive item variable into (item_code, column, type, unit)."""
    for suffix, column, bq_type, unit in ITEM_MEASURES:
        if len(name) > len(suffix) and name.endswith(suffix):
            return name[: -len(suffix)], column, bq_type, unit
    return None


def sas_code_as_written(code: str, cycle: str) -> str:
    """How a SAS special-missing code actually appears in the CSV Public Use File.

    Neither cycle writes the SPSS numeric codes the codebook also lists. Cycle 1
    strips the leading dot and upper-cases (.N -> N); Cycle 2 keeps the dot and
    lower-cases (.N -> .n). Verified against JPN, DEU, ECU and the US national
    file for Cycle 1, and AUT and DEU for Cycle 2.
    """
    code = code.strip()
    if not code.startswith("."):
        return code
    if code == ".":
        return "."
    return code.lstrip(".").upper() if cycle == "1" else code.lower()


def sas_code_variants(code: str) -> set[str]:
    """Every spelling of a SAS missing code, so a match never depends on cycle.

    Safe to over-accept: these are letter tokens, and the columns they are
    stripped from hold numbers. Legitimate single-letter values do exist in the
    data -- ISIC section codes run A to U -- which is exactly why the accepted
    set is built per column from the codebook rather than from a global token list.
    """
    code = code.strip()
    if not code or code == ".":
        return {"."} if code else set()
    bare = code.lstrip(".")
    return {code, code.lower(), code.upper(), bare, bare.lower(), bare.upper()}


class UnclassifiedColumnError(ValueError):
    """Raised when no rule assigns a type, rather than guessing one."""


def classify(variable: Variable) -> tuple[str, str, str, str]:
    """Return (bigquery_type, covered_by_dictionary, measurement_unit, observations).

    A coded value set always wins: PIAAC's reserved codes (6 valid skip, 7 don't
    know, 8 refused, 9 not stated, and the 96-99 / 9996-9999 variants) are
    meaningful answers, and casting them to a number would turn "refused" into a
    quantity.
    """
    if variable.has_value_scheme:
        return "STRING", "yes", "", ""

    if STRING_BY_NAME.match(variable.name):
        return "STRING", "no", "", ""

    for pattern, unit, explanation in UNITLESS_NUMERIC:
        if pattern.search(variable.name):
            return "FLOAT64", "no", unit, explanation

    for pattern, unit, explanation in UNITLESS_NUMERIC_BY_LABEL:
        if pattern.search(variable.label):
            return "FLOAT64", "no", unit, explanation

    measure = MEASURE_PARENTHETICAL.search(variable.label)
    if measure:
        unit = {
            "number of actions": "action",
            "timing": "second",
            "timing first action": "second",
            "visit": "visit",
            "short visit": "visit",
        }[measure.group(1).strip().lower()]
        return ("FLOAT64" if variable.decimals else "INT64"), "no", unit, ""

    haystack = f"{variable.name} {variable.label}"
    for pattern, unit in UNIT_RULES:
        if pattern.search(haystack):
            bq_type = "FLOAT64" if variable.decimals else "INT64"
            observations = ""
            if not unit:
                observations = (
                    "Reported in the respondent's national currency, so the unit "
                    "varies by country"
                    if re.search(
                        r"earn|pay|income|wage|salary", haystack, re.I
                    )
                    else "Simple count; the counted item is named in the description"
                )
            return bq_type, "no", unit, observations

    if variable.level in {"Ratio", "Scale", "Interval"}:
        raise UnclassifiedColumnError(
            f"cycle {variable.cycle} {variable.name!r} is {variable.level} with no "
            f"value scheme and no unit rule: {variable.label!r}"
        )

    return "STRING", "no", "", ""
