"""Generate the architecture CSVs and the codebook-derived reference tables.

Usage:
    uv run python models/world_oecd_piaac/code/build_architecture.py

Writes:
    models/world_oecd_piaac/code/architecture/<table>.csv   schema definitions
    <output>/variable/data.parquet                          the codebook as a table
    <output>/dictionary/data.parquet                        coded value -> meaning

The architecture CSVs are the source of truth for the cleaning transform, the dbt
models and the backend column registration, so all three stay consistent.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from dataclasses import asdict
from pathlib import Path

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq

sys.path.insert(0, str(Path(__file__).parent))

import architecture as arch
import codebook as cb
import constants as piaac

ARCHITECTURE_DIR = Path(__file__).parent / "architecture"
CSV_FIELDS = [
    "name",
    "bigquery_type",
    "description",
    "temporal_coverage",
    "covered_by_dictionary",
    "directory_column",
    "measurement_unit",
    "has_sensitive_data",
    "observations",
    "original_name",
]

# Value schemes that name a lookup sheet instead of listing pairs. Occupation and
# industry resolve to directories rather than this dataset's dictionary, per the
# shared-entity rule in .claude/rules/data-basis-style.md.
SHEET_DIRECTORY = {
    "ISCO08": "br_bd_diretorios_mundo.isco_08:id_isco_08",
    "ISCO08_2digit": "br_bd_diretorios_mundo.isco_08:id_isco_08",
    "ISIC4": "br_bd_diretorios_mundo.isic_4:id_isic_4",
    "ISIC4_2digit": "br_bd_diretorios_mundo.isic_4:id_isic_4",
}
# Cycle 1 names its occupation and industry variables plainly. ISCOSKIL4 is
# deliberately NOT here: despite the name it is a four-category skill grouping,
# not an ISCO code, and its values 1-4 would collide with ISCO major groups 1-4.
CY1_DIRECTORY = [
    (
        re.compile(r"^ISCO[12][CL]$", re.I),
        "br_bd_diretorios_mundo.isco_08:id_isco_08",
    ),
    (
        re.compile(r"^ISIC[12][CL]$", re.I),
        "br_bd_diretorios_mundo.isic_4:id_isic_4",
    ),
]

TEMPORAL_COVERAGE = {"1": "2012(1)2017", "2": "2023(1)2023"}
RENAMED_INTO_GRAIN_LOWER = {"cntryid_e": "country_entity_id"}


def split_packed(scheme: str) -> list[tuple[str, str]]:
    """Split "1: Male; 2: Female" into pairs.

    Labels contain semicolons -- ISIC section names such as "Wholesale and retail
    trade; repair of motor vehicles" -- so a fragment with no colon is a
    continuation of the previous pair rather than a new one.
    """
    pairs, current = [], None
    for chunk in scheme.split(";"):
        if ":" in chunk:
            if current is not None:
                pairs.append(current)
            current = chunk
        elif current is not None:
            current += ";" + chunk
    if current is not None:
        pairs.append(current)
    out = []
    for pair in pairs:
        key, _, value = pair.partition(":")
        key, value = key.strip(), value.strip()
        if key and value:
            out.append((key, value))
    return out


def load_lookup_sheets(path: Path) -> dict[str, list[tuple[str, str]]]:
    """The Cycle 2 codebook holds its largest value sets as separate sheets."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in workbook.sheetnames:
        if name == "PUF":
            continue
        pairs = []
        for row in workbook[name].iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            pairs.append((str(row[0]).strip(), str(row[1]).strip()))
        sheets[name] = pairs
    workbook.close()
    return sheets


def directory_for(variable: cb.Variable, sheets: dict) -> str:
    scheme = variable.value_scheme.strip()
    if scheme in SHEET_DIRECTORY:
        return SHEET_DIRECTORY[scheme]
    if variable.cycle == "1":
        for pattern, directory in CY1_DIRECTORY:
            if pattern.match(variable.name):
                return directory
    return ""


RENAMED_INTO_GRAIN = {"CNTRYID_E": "country_entity_id"}


def respondent_columns(
    variables: list[cb.Variable], cycle: str, sheets: dict
) -> list[arch.Column]:
    """Grain columns, then every non-item variable under its own PIAAC name."""
    skip = {"SEQID", "CNTRYID", "CNTRYID_E", "CNTRY", "CNTRY_E"}
    columns = arch.grain_columns(cycle)
    # CNTRYID_E is a coded list of countries and sub-national entities, and it
    # keeps that coding under its new name.
    for column in columns:
        if column.name == "country_entity_id":
            column.covered_by_dictionary = "yes"
    for variable in variables:
        if variable.is_item or variable.name.upper() in skip:
            continue
        bq_type, covered, unit, observations = cb.classify(variable)
        directory = directory_for(variable, sheets)
        if directory:
            # A directory is the source of truth for a shared entity, so the column
            # is not covered by this dataset's dictionary.
            covered = "no"
        columns.append(
            arch.Column(
                name=variable.name.lower(),
                bigquery_type=bq_type,
                description=variable.label or variable.name,
                covered_by_dictionary=covered,
                directory_column=directory,
                measurement_unit=unit,
                observations=observations,
                original_name=variable.name,
            )
        )
    return columns


def usa_national_columns(
    header: list[str], cy1: list[cb.Variable], sheets: dict
) -> list[arch.Column]:
    """The US national file shares most variables with the international codebook.

    It suppresses 110 international variables and adds 131 US-only ones that the
    international codebook does not describe; those are labelled honestly rather
    than given an invented description.
    """
    by_upper = {v.name.upper(): v for v in cy1}
    skip = {"SEQID", "CNTRYID", "CNTRYID_E", "CNTRY", "CNTRY_E"}
    columns = arch.grain_columns("1")
    for raw in header:
        upper = raw.upper()
        if upper in skip:
            continue
        variable = by_upper.get(upper)
        if variable is None:
            columns.append(
                arch.Column(
                    name=raw.lower(),
                    bigquery_type="STRING",
                    description=f"United States national variable {raw}",
                    observations=(
                        "Not present in the PIAAC international codebook. See the "
                        "U.S. PIAAC national codebook for its definition"
                    ),
                    original_name=raw,
                )
            )
            continue
        if variable.is_item:
            continue
        bq_type, covered, unit, observations = cb.classify(variable)
        directory = directory_for(variable, sheets)
        columns.append(
            arch.Column(
                name=raw.lower(),
                bigquery_type=bq_type,
                description=variable.label or raw,
                covered_by_dictionary="no" if directory else covered,
                directory_column=directory,
                measurement_unit=unit,
                observations=observations,
                original_name=raw,
            )
        )
    return columns


def write_architecture(
    slug: str, columns: list[arch.Column], cycle: str | None
) -> None:
    ARCHITECTURE_DIR.mkdir(parents=True, exist_ok=True)
    with (ARCHITECTURE_DIR / f"{slug}.csv").open(
        "w", newline="", encoding="utf-8"
    ) as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for column in columns:
            row = asdict(column)
            if cycle and not row["temporal_coverage"]:
                row["temporal_coverage"] = ""
            writer.writerow({k: row[k] for k in CSV_FIELDS})


def write_parquet(records: list[dict], slug: str) -> int:
    columns = list(records[0])
    table = pa.table(
        {
            name: pa.array([r[name] for r in records], type=pa.string())
            for name in columns
        }
    )
    destination = piaac.OUTPUT_ROOT / slug
    destination.mkdir(parents=True, exist_ok=True)
    pq.write_table(table, destination / "data.parquet", compression="snappy")
    return len(records)


def reserved_codes(
    records: list[dict], numeric_by_table: dict[str, set[str]]
) -> dict:
    """Codes on numeric columns that mean "no answer", not a quantity.

    PIAAC pads reserved codes to the width of the field, so hourly earnings carry
    999999999996 for "valid skip". Casting those to a number without nulling them
    first would put a trillion-dollar wage into every mean, so the cleaning
    transform nulls exactly these values and each column records which ones.
    """
    codes: dict[str, dict[str, dict[str, str]]] = {}
    for record in records:
        table_id, column = record["table_id"], record["column_name"]
        if column not in numeric_by_table.get(table_id, set()):
            continue
        # Accept every spelling of the code, not only this cycle's canonical one,
        # so a country file that deviates cannot slip a sentinel into a numeric
        # column. Safe to over-accept: these are letter tokens and the columns
        # they are stripped from hold numbers.
        target = codes.setdefault(table_id, {}).setdefault(column, {})
        for variant in cb.sas_code_variants(record["key"]) or {record["key"]}:
            target[variant] = record["value"]
    return codes


def build_dictionary(all_variables: dict, sheets: dict) -> list[dict]:
    """Coded value -> meaning, for dictionary-covered respondent columns only.

    Item columns are excluded on purpose: their codes are item-specific, so the
    decoded meaning travels inline as item_response.scored_response_label.
    """
    records, seen = [], set()
    for table_id, (variables, cycle) in all_variables.items():
        for variable in variables:
            if variable.is_item:
                continue
            if directory_for(variable, sheets):
                continue
            # A variable with no value scheme is a quantity, but it can still carry
            # reserved codes in its missing scheme -- and those must be collected,
            # or the cleaning transform would leave 9999 sitting in a numeric
            # column. They are filtered back out of the published dictionary.
            if (
                not variable.has_value_scheme
                and not variable.missing_scheme_sas
            ):
                continue
            scheme = variable.value_scheme.strip()
            pairs = sheets.get(scheme) or split_packed(scheme)
            # Reserved codes are answers too -- refused, don't know, valid skip --
            # and belong in the dictionary alongside the substantive ones. Key
            # them on the SAS coding, which is what the CSV PUFs carry; the SPSS
            # numeric codes the codebook lists beside them appear nowhere in the
            # data.
            if variable.missing_scheme_sas:
                pairs = pairs + [
                    (cb.sas_code_as_written(k, variable.cycle), v)
                    for k, v in split_packed(variable.missing_scheme_sas)
                    if k.strip() and k.strip() != "."
                ]
            # The data carries the numeric reserved family on columns whose
            # codebook entry only declares SAS letters. Existing pairs win.
            declared = {k for k, _ in pairs}
            pairs = pairs + [
                (k, v)
                for k, v in cb.reserved_code_family(variable.width).items()
                if k not in declared
            ]
            column = variable.name.lower()
            for key, value in pairs:
                identity = (table_id, column, key)
                if identity in seen:
                    continue
                seen.add(identity)
                records.append(
                    {
                        "table_id": table_id,
                        "column_name": column,
                        "key": key,
                        "temporal_coverage": TEMPORAL_COVERAGE[cycle],
                        "value": value,
                    }
                )
    return records


def build_cycle_1_dictionary(
    path: Path,
    respondent_names: set[str],
    sheets: dict,
    widths: dict[str, int],
) -> list[dict]:
    """Cycle 1 publishes its value labels as a long sheet rather than packed text."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records, seen = [], set()
    for row in workbook["Values"].iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[3] is None:
            continue
        column = str(row[0]).strip().lower()
        if column not in respondent_names:
            continue
        # Column 2 is the SAS coding, which is what the CSV PUFs actually carry:
        # Cycle 1 writes .N as a bare N. Column 3 is the SPSS numeric coding,
        # which appears nowhere in the CSVs. Valid values are identical in both
        # codings (verified: zero disagreements), so SAS is safe throughout.
        raw_key = str(row[2]).strip()
        if not raw_key or raw_key in {"None", "."}:
            continue
        key, value = cb.sas_code_as_written(raw_key, "1"), str(row[1]).strip()
        identity = ("respondent_cycle_1", column, key)
        if identity in seen:
            continue
        seen.add(identity)
        records.append(
            {
                "table_id": "respondent_cycle_1",
                "column_name": column,
                "key": key,
                "temporal_coverage": TEMPORAL_COVERAGE["1"],
                "value": value,
            }
        )
    workbook.close()

    # As in Cycle 2: the coded language, occupation and industry columns carry the
    # numeric reserved family in the data even though the Values sheet lists only
    # the SAS letters for them.
    declared = {(r["column_name"], r["key"]) for r in records}
    for column in sorted({r["column_name"] for r in records}):
        for key, value in cb.reserved_code_family(
            widths.get(column, 0)
        ).items():
            if (column, key) not in declared:
                records.append(
                    {
                        "table_id": "respondent_cycle_1",
                        "column_name": column,
                        "key": key,
                        "temporal_coverage": TEMPORAL_COVERAGE["1"],
                        "value": value,
                    }
                )
    return records


def main() -> None:
    docs = piaac.DOCS_ROOT
    cy1 = cb.load_codebook(
        docs / "cycle_1" / "international_codebook.xlsx", "1"
    )
    cy2 = cb.load_codebook(
        docs / "cycle_2" / "international_codebook.xlsx", "2"
    )
    sheets = load_lookup_sheets(
        docs / "cycle_2" / "international_codebook.xlsx"
    )

    tables: dict[str, list[arch.Column]] = {
        "respondent_cycle_1": respondent_columns(cy1, "1", sheets),
        "respondent_cycle_2": respondent_columns(cy2, "2", sheets),
        "item_response_cycle_1": arch.grain_columns("1") + arch.ITEM_COLUMNS,
        "item_response_cycle_2": arch.grain_columns("2") + arch.ITEM_COLUMNS,
        "variable": arch.VARIABLE_COLUMNS,
        "dictionary": arch.DICTIONARY_COLUMNS,
    }

    usa_path = piaac.local_puf_path("USA", "1", "3", "Prgusap1_2017.csv")
    if usa_path.exists():
        header = (
            usa_path.open(encoding="utf-8", errors="replace")
            .readline()
            .strip()
        )
        tables["respondent_cycle_1_usa_national"] = usa_national_columns(
            [c.strip('"') for c in header.split("|")], cy1, sheets
        )
    else:
        print(
            f"  [skip] respondent_cycle_1_usa_national: {usa_path} not downloaded yet"
        )

    for slug, columns in tables.items():
        write_architecture(slug, columns, None)
        print(f"  {slug:<34} {len(columns):>5} columns")

    # --- reference tables -------------------------------------------------
    variable_records = []
    for cycle, variables in (("1", cy1), ("2", cy2)):
        for variable in variables:
            item = cb.split_item(variable.name) if variable.is_item else None
            if item:
                table_id, column_name = f"item_response_cycle_{cycle}", item[1]
                bq_type, unit = item[2], item[3]
            else:
                table_id, column_name = (
                    f"respondent_cycle_{cycle}",
                    variable.name.lower(),
                )
                bq_type, _, unit, _ = cb.classify(variable)
            variable_records.append(
                {
                    "cycle": cycle,
                    "variable_name": variable.name.lower(),
                    "table_id": table_id,
                    "column_name": column_name,
                    "label": variable.label,
                    "domain": variable.domain,
                    "level": variable.level,
                    "bigquery_type": bq_type,
                    "measurement_unit": unit,
                    "item_code": item[0].lower() if item else "",
                    "measure": item[1] if item else "",
                }
            )
    print(
        f"  variable table                     {write_parquet(variable_records, 'variable'):>5} rows"
    )

    respondent_1 = {v.name.lower() for v in cy1 if not v.is_item}
    widths_1 = {v.name.lower(): v.width for v in cy1}
    dictionary = build_cycle_1_dictionary(
        docs / "cycle_1" / "international_codebook.xlsx",
        respondent_1,
        sheets,
        widths_1,
    )
    dictionary += build_dictionary({"respondent_cycle_2": (cy2, "2")}, sheets)

    # CNTRYID_E's value set follows it into its renamed column.
    for record in dictionary:
        if record["column_name"] in RENAMED_INTO_GRAIN_LOWER:
            record["column_name"] = RENAMED_INTO_GRAIN_LOWER[
                record["column_name"]
            ]

    numeric_by_table, declared_by_table = {}, {}
    for slug, columns in tables.items():
        numeric_by_table[slug] = {
            c.name for c in columns if c.bigquery_type in {"INT64", "FLOAT64"}
        }
        declared_by_table[slug] = {
            c.name for c in columns if c.covered_by_dictionary == "yes"
        }

    # The US national table shares Cycle 1's coded columns, but the coverage test
    # matches on table_id, so it needs its own copy of the relevant entries.
    usa_covered = declared_by_table.get(
        "respondent_cycle_1_usa_national", set()
    )
    dictionary += [
        {**record, "table_id": "respondent_cycle_1_usa_national"}
        for record in list(dictionary)
        if record["table_id"] == "respondent_cycle_1"
        and record["column_name"] in usa_covered
    ]

    codes = reserved_codes(dictionary, numeric_by_table)
    (ARCHITECTURE_DIR / "reserved_codes.json").write_text(
        json.dumps(codes, indent=2, sort_keys=True, ensure_ascii=False),
        encoding="utf-8",
    )
    n_codes = sum(len(v) for t in codes.values() for v in t.values())
    n_columns = sum(len(t) for t in codes.values())
    print(
        f"  reserved codes                     {n_codes:>5} across {n_columns} numeric columns"
    )

    # Record on each numeric column exactly which values the cleaning transform
    # nulls, so the loss is documented rather than silent.
    for slug, columns in tables.items():
        for column in columns:
            entry = codes.get(slug, {}).get(column.name)
            if not entry:
                continue
            listed = ", ".join(sorted({f"{v}" for v in entry.values()}))
            note = (
                f"Set to NULL when loading where the source recorded: {listed}"
            )
            column.observations = (
                f"{column.observations}. {note}"
                if column.observations
                else note
            )
        write_architecture(slug, columns, None)

    # A dictionary row for a column that is not dictionary-covered would be dead
    # weight -- and misleading on numeric columns, whose codes are nulled on load.
    dictionary = [
        r
        for r in dictionary
        if r["column_name"] in declared_by_table.get(r["table_id"], set())
    ]
    print(
        f"  dictionary table                   {write_parquet(dictionary, 'dictionary'):>5} rows"
    )


if __name__ == "__main__":
    main()
