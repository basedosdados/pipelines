"""Universal parser for the ABS Australian System of National Accounts (5204.0).

Every ASNA time-series workbook shares one layout: an ``Index`` sheet plus one or
more ``Data`` sheets carrying a fixed metadata block (Unit, Series Type, Data
Type, Frequency, Collection Month, Series Start/End, No. Obs, Series ID) above
June-dated annual rows. This single parser ingests all 71 standard files into a
long two-table design keyed on the ABS Series ID:

  - series.parquet              one row per series_id (the dictionary/dimension)
  - observations/year=YYYY/...  long fact (year, financial_year, series_id, value)

Table 14 (Market Sector Productivity) is skipped: it is a growth-cycle summary
with no Series IDs, not an annual time series.

Usage:
    python clean_data.py <input_dir_with_xlsx> <output_dir>
"""

import datetime as dt
import glob
import os
import sys

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

SKIP_FILES = {"5204014_Market_Sector_Productivity"}

# Series Type (always "Original"), Data Type ("DERIVED") and Frequency ("Annual")
# are constant across all ~5,190 ASNA series, so they are dropped rather than
# stored. Only these labels are needed to parse the block.
META_LABELS = {
    "Unit": "unit",
    "Series Start": "series_start",
    "Series End": "series_end",
    "Series ID": "series_id",
}


def _table_no_name(base: str):
    """`5204005_GVA_by_Industry` -> ("5", "GVA by Industry")."""
    stem, _, rest = base.partition("_")
    table_no = str(int(stem[4:7]))  # 5204005 -> 005 -> 5
    table_name = rest.replace("_", " ").strip()
    return table_no, table_name


def _fin_year(end_year: int) -> str:
    """Calendar year the FY ends in -> "YYYY-YY" (2025 -> "2024-25")."""
    return f"{end_year - 1}-{end_year % 100:02d}"


def parse_workbook(path: str):
    base = os.path.basename(path).replace(".xlsx", "")
    table_no, table_name = _table_no_name(base)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data_sheets = [s for s in wb.sheetnames if s.lower().startswith("data")]

    series_rows = []
    value_rows = []
    n_nonnumeric = 0

    for sheet in data_sheets:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Locate the metadata block by the label in column A.
        label_idx = {}
        for i, r in enumerate(rows[:15]):
            a = r[0]
            if isinstance(a, str):
                a = a.strip().rstrip(".")
                for lbl in META_LABELS:
                    if a == lbl.rstrip("."):
                        label_idx[lbl] = i
        if "Series ID" not in label_idx:
            continue

        desc_row = rows[0]
        sid_row = rows[label_idx["Series ID"]]
        unit_row = rows[label_idx["Unit"]]
        sstart_row = rows[label_idx["Series Start"]]
        send_row = rows[label_idx["Series End"]]
        data_start = max(label_idx.values()) + 1

        # Valid series columns are those with a Series ID.
        cols = [
            j for j in range(1, len(sid_row)) if sid_row[j] not in (None, "")
        ]

        def _d(v):
            return v.date() if isinstance(v, dt.datetime) else None

        def _desc(v):
            # ABS descriptions carry a trailing " ;" separator artifact; strip it.
            if v is None:
                return None
            return str(v).strip().rstrip(";").strip() or None

        for j in cols:
            series_rows.append(
                {
                    "series_id": str(sid_row[j]).strip(),
                    "description": _desc(desc_row[j]),
                    "unit": (
                        str(unit_row[j]).strip()
                        if unit_row[j] is not None
                        else None
                    ),
                    "source_table_no": table_no,
                    "source_table_name": table_name,
                    "series_start": _d(sstart_row[j]),
                    "series_end": _d(send_row[j]),
                }
            )

        for r in rows[data_start:]:
            d = r[0]
            if not isinstance(d, dt.datetime):
                continue
            year = d.year
            fy = _fin_year(year)
            for j in cols:
                v = r[j]
                if v is None or v == "":
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    n_nonnumeric += 1
                    continue
                value_rows.append(
                    {
                        "year": year,
                        "financial_year": fy,
                        "series_id": str(sid_row[j]).strip(),
                        "value": fv,
                    }
                )

    wb.close()
    return series_rows, value_rows, n_nonnumeric


def main(input_dir: str, output_dir: str):
    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    files = [
        f
        for f in files
        if os.path.basename(f).replace(".xlsx", "") not in SKIP_FILES
    ]
    print(f"Parsing {len(files)} files (Table 14 skipped)")

    all_series, all_values, nonnum = [], [], 0
    for f in files:
        s, v, nn = parse_workbook(f)
        all_series.extend(s)
        all_values.extend(v)
        nonnum += nn

    series = pd.DataFrame(all_series)
    values = pd.DataFrame(all_values)

    # ---- series dictionary: one row per series_id (first occurrence wins) ----
    dup_series = series["series_id"].duplicated().sum()
    series = series.drop_duplicates(
        subset="series_id", keep="first"
    ).reset_index(drop=True)

    # ---- values: dedup cross-listed (year, series_id); flag any conflicts ----
    before = len(values)
    conflicts = values.groupby(["year", "series_id"])["value"].nunique()
    n_conflict = int((conflicts > 1).sum())
    values = values.drop_duplicates(
        subset=["year", "series_id"], keep="first"
    ).reset_index(drop=True)

    print(
        f"series rows: {len(series)}  (dropped {dup_series} cross-listed duplicates)"
    )
    print(
        f"values rows: {len(values)}  (dropped {before - len(values)} cross-listed duplicates)"
    )
    print(f"value cells that were non-numeric and skipped: {nonnum}")
    print(
        f"(year, series_id) pairs with CONFLICTING values across tables: {n_conflict}"
    )
    print(f"year range: {values['year'].min()}..{values['year'].max()}")
    print(f"distinct units: {sorted(series['unit'].dropna().unique())}")

    os.makedirs(output_dir, exist_ok=True)

    # series.parquet (explicit schema; DATE for the two date columns)
    series = series[
        [
            "series_id",
            "description",
            "unit",
            "source_table_no",
            "source_table_name",
            "series_start",
            "series_end",
        ]
    ]
    series_schema = pa.schema(
        [
            ("series_id", pa.string()),
            ("description", pa.string()),
            ("unit", pa.string()),
            ("source_table_no", pa.string()),
            ("source_table_name", pa.string()),
            ("series_start", pa.date32()),
            ("series_end", pa.date32()),
        ]
    )
    series_tbl = pa.Table.from_pandas(
        series, schema=series_schema, preserve_index=False
    )
    series_dir = os.path.join(output_dir, "series")
    os.makedirs(series_dir, exist_ok=True)
    pq.write_table(
        series_tbl,
        os.path.join(series_dir, "series.parquet"),
        compression="snappy",
    )

    # values: partitioned by year
    values = values[["year", "financial_year", "series_id", "value"]]
    values_schema = pa.schema(
        [
            ("year", pa.int64()),
            ("financial_year", pa.string()),
            ("series_id", pa.string()),
            ("value", pa.float64()),
        ]
    )
    values_tbl = pa.Table.from_pandas(
        values, schema=values_schema, preserve_index=False
    )
    values_dir = os.path.join(output_dir, "observations")
    pq.write_to_dataset(
        values_tbl,
        root_path=values_dir,
        partition_cols=["year"],
        compression="snappy",
    )
    print(f"\nWrote {series_dir}/series.parquet and {values_dir}/year=*/")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
