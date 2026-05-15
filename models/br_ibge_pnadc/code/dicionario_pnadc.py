"""
Script Unificado de Processamento do Dicionário PNADC
======================================================

Este script processa dois arquivos Excel:
1. dicionario_PNADC_microdados_trimestral.xls - Dicionário PNADC
2. Estrutura_Ocupacao_COD.xlsx - Estrutura de Ocupação

E gera um único arquivo Parquet com todas as transformações:
- Renomeação de colunas
- Classificação de registros (educacao/microdados)
- Integração de dados de ocupação (V4010)
- Preservação de zeros à esquerda
- Preenchimento de cobertura_temporal com "(1)"

Autor: Data Engineering Team
Data: 2026-05-13
"""

import platform
import sys
from pathlib import Path

import openpyxl
import pandas as pd

# ===== CONFIGURAÇÕES =====
print("=" * 80)
print("PROCESSAMENTO UNIFICADO DO DICIONÁRIO PNADC")
print("=" * 80)

# Detectar sistema operacional e ajustar caminhos

is_windows = platform.system() == "Windows"

if is_windows:
    # Caminho para Windows
    base_path = Path(r"D:\Devops\Git\BD\areadetrabalho")
    arquivo_pnadc = base_path / "dicionario_PNADC_microdados_trimestral.xls"
    arquivo_ocupacao = base_path / "Estrutura_Ocupacao_COD.xlsx"
    arquivo_saida = base_path / "dicionario_dbt_final.parquet"
else:
    # Caminho para Linux/Unix (para testes)
    base_path = Path("/home/ubuntu/upload")
    arquivo_pnadc = base_path / "dicionario_PNADC_microdados_trimestral.xls"
    arquivo_ocupacao = base_path / "Estrutura_Ocupacao_COD.xlsx"
    arquivo_saida = Path("/home/ubuntu/output/dicionario_dbt_final.parquet")

print(f"\nSistema Operacional: {platform.system()}")
print(f"Caminho Base: {base_path}")
print(f"Arquivo PNADC: {arquivo_pnadc}")
print(f"Arquivo Ocupação: {arquivo_ocupacao}")
print(f"Arquivo Saída: {arquivo_saida}")

# ===== ETAPA 1: Leitura do Dicionário PNADC =====
print("\n" + "=" * 80)
print("ETAPA 1: Lendo dicionário PNADC...")
print("=" * 80)

try:
    # Ler sem header para processar manualmente
    df_pnadc_raw = pd.read_excel(arquivo_pnadc, sheet_name=0, header=None)
    print("✓ Arquivo PNADC lido")
    print(f"Shape: {df_pnadc_raw.shape}")
except Exception as e:
    print(f"✗ Erro ao ler arquivo PNADC: {e}")
    sys.exit(1)

# ===== ETAPA 2: Processar Dicionário PNADC =====
print("\n" + "=" * 80)
print("ETAPA 2: Processando dicionário PNADC...")
print("=" * 80)

# Estrutura do arquivo PNADC:
# Col0: Posição inicial
# Col1: Tamanho
# Col2: Código da variável (NOME_COLUNA)
# Col3: Código adicional
# Col4: Descrição da variável
# Col5: Tipo/Chave (CHAVE)
# Col6: Valor/Descrição (VALOR)
# Col7: Período

print("Mapeamento de colunas:")
print("  Col2: Código da variável (NOME_COLUNA)")
print("  Col5: Tipo/Chave (CHAVE)")
print("  Col6: Valor/Descrição (VALOR)")

registros_pnadc = []
id_tabela_atual = "microdados"

for _, row in df_pnadc_raw.iterrows():
    # Pegar valores das colunas conforme mapeamento correto
    col0 = (
        str(row[0]).strip() if pd.notna(row[0]) else ""
    )  # Posição inicial (ou "Parte")
    col2 = (
        str(row[2]).strip() if pd.notna(row[2]) else ""
    )  # Código da variável (NOME_COLUNA)
    col5 = (
        str(row[5]).strip() if pd.notna(row[5]) else ""
    )  # Tipo/Chave (CHAVE)
    col6 = (
        str(row[6]).strip() if pd.notna(row[6]) else ""
    )  # Valor/Descrição (VALOR)

    # Verificar se é uma linha de "Parte" (em Col0)
    if "Parte" in col0:
        if "Características de educação" in col0:
            id_tabela_atual = "educacao"
        else:
            id_tabela_atual = "microdados"
        continue

    # Pular linhas de header ou vazias
    if col2 in [
        "Dicionário das variáveis da PNAD contínua",
        "Posição inicial",
        "Código da variável",
        "Tipo",
        "Descrição",
    ]:
        continue

    if not col2 or col2 == "NaN":
        continue

    # Se temos código e valor válidos
    if (
        col2
        and col6
        and col6.lower() not in ["não aplicável", "nao aplicavel"]
    ):
        registros_pnadc.append(
            {
                "id_tabela": id_tabela_atual,
                "nome_coluna": col2,
                "chave": col5,
                "cobertura_temporal": "(1)",
                "valor": col6,
            }
        )

df_pnadc = pd.DataFrame(registros_pnadc)
print(f"\nTotal de registros PNADC: {len(df_pnadc)}")

if len(df_pnadc) > 0:
    print(
        f"Registros de educacao: {(df_pnadc['id_tabela'] == 'educacao').sum()}"
    )
    print(
        f"Registros de microdados: {(df_pnadc['id_tabela'] == 'microdados').sum()}"
    )

    # Mostrar alguns exemplos
    print("\nExemplos de registros PNADC:")
    print(df_pnadc.head(10).to_string())

    # Remover V4010 do PNADC (será substituído pelos dados do Excel)
    df_pnadc = (
        df_pnadc[df_pnadc["nome_coluna"] != "V4010"]
        .copy()
        .reset_index(drop=True)
    )
    print(f"\nTotal após remover V4010: {len(df_pnadc)}")
else:
    print("⚠️ Nenhum registro PNADC processado!")

# ===== ETAPA 3: Leitura e Processamento da Estrutura de Ocupação =====
print("\n" + "=" * 80)
print("ETAPA 3: Lendo e processando estrutura de ocupação...")
print("=" * 80)

try:
    wb = openpyxl.load_workbook(arquivo_ocupacao)
    ws = wb.active
    print("✓ Arquivo de ocupação lido")
    print(f"Total de linhas: {ws.max_row}")
except Exception as e:
    print(f"✗ Erro ao ler arquivo de ocupação: {e}")
    sys.exit(1)

# Processar dados de ocupação
registros_v4010 = []
chaves_processadas = set()

for row_idx in range(3, ws.max_row + 1):
    gg = ws[f"A{row_idx}"].value
    sg = ws[f"B{row_idx}"].value
    sub = ws[f"C{row_idx}"].value
    gb = ws[f"D{row_idx}"].value
    den = ws[f"E{row_idx}"].value

    # Determinar qual coluna tem valor (prioridade: GB > SUB > SG > GG)
    chave = None
    if gb is not None:
        chave = str(gb)
    elif sub is not None:
        chave = str(sub)
    elif sg is not None:
        chave = str(sg)
    elif gg is not None:
        chave = str(gg)

    # Se temos uma chave e denominação válida
    if chave and den is not None:
        den_str = str(den).strip().lower()

        # Combinando as verificações de conteúdo e de duplicata em um único if
        if (
            den_str not in ["nao aplicavel", "não aplicável"]
            and "codigo" not in den_str
            and chave not in chaves_processadas
        ):
            chaves_processadas.add(chave)

            registros_v4010.append(
                {
                    "id_tabela": "microdados",
                    "nome_coluna": "V4010",
                    "chave": chave,
                    "cobertura_temporal": "(1)",
                    "valor": str(den).strip(),
                }
            )

df_v4010 = pd.DataFrame(registros_v4010)

# Remover registros com chave vazia ou NaN ou com valor "Não aplicável"
df_v4010 = df_v4010[df_v4010["chave"].notna()].copy()
df_v4010 = df_v4010[df_v4010["chave"].astype(str).str.strip() != ""].copy()
df_v4010 = df_v4010[
    ~df_v4010["valor"]
    .astype(str)
    .str.lower()
    .isin(["nao aplicavel", "não aplicável"])
].copy()
df_v4010 = df_v4010.reset_index(drop=True)

print(f"Total de registros V4010: {len(df_v4010)}")
print(f"Chaves únicas: {df_v4010['chave'].nunique()}")

# ===== ETAPA 4: Combinar PNADC com V4010 =====
print("\n" + "=" * 80)
print("ETAPA 4: Combinando PNADC com dados de ocupação...")
print("=" * 80)

df_final = pd.concat([df_pnadc, df_v4010], ignore_index=True)

print(f"Total de registros finais: {len(df_final)}")
print(f"Registros de educacao: {(df_final['id_tabela'] == 'educacao').sum()}")
print(
    f"Registros de microdados: {(df_final['id_tabela'] == 'microdados').sum()}"
)

# ===== ETAPA 5: Ordenação dos dados =====
print("\n" + "=" * 80)
print("ETAPA 5: Ordenando dados...")
print("=" * 80)

df_final = df_final.sort_values(
    by=["id_tabela", "nome_coluna", "chave"]
).reset_index(drop=True)
print("✓ Dados ordenados por: id_tabela, nome_coluna, chave")

# ===== ETAPA 6: Garantir tipos de dados =====
print("\n" + "=" * 80)
print("ETAPA 6: Garantindo tipos de dados...")
print("=" * 80)

df_final = df_final.astype(
    {
        "id_tabela": "str",
        "nome_coluna": "str",
        "chave": "str",
        "cobertura_temporal": "str",
        "valor": "str",
    }
)

print("Tipos de dados:")
print(df_final.dtypes)

# ===== ETAPA 7: Verificação de duplicatas =====
print("\n" + "=" * 80)
print("ETAPA 7: Verificação de duplicatas...")
print("=" * 80)

v4010_final = df_final[df_final["nome_coluna"] == "V4010"]
print(f"Total de registros V4010: {len(v4010_final)}")
print(f"Chaves únicas: {v4010_final['chave'].nunique()}")

if len(v4010_final) > 0 and v4010_final["chave"].nunique() == len(v4010_final):
    print("✓ SEM DUPLICATAS EM V4010!")
else:
    print("⚠️ Verificar duplicatas")

# Verificar zeros à esquerda
if len(v4010_final) > 0:
    v4010_com_zeros = v4010_final[
        v4010_final["chave"].astype(str).str.startswith("0")
    ]
    print("\nExemplos de V4010 com zeros à esquerda:")
    print(f"Total com zeros à esquerda: {len(v4010_com_zeros)}")
    if len(v4010_com_zeros) > 0:
        print(
            v4010_com_zeros[["nome_coluna", "chave", "valor"]]
            .head(15)
            .to_string()
        )

# ===== ETAPA 8: Exportando dados =====
print("\n" + "=" * 80)
print("ETAPA 8: Exportando dados...")
print("=" * 80)

# Criar diretório de saída se não existir
arquivo_saida.parent.mkdir(parents=True, exist_ok=True)

# Exportar para Parquet
try:
    df_final.to_parquet(arquivo_saida, index=False, compression="snappy")
    print(f"✓ Exportado para Parquet: {arquivo_saida}")
except Exception as e:
    print(f"✗ Erro ao exportar Parquet: {e}")
    sys.exit(1)

# ===== ETAPA 9: Verificação Final =====
print("\n" + "=" * 80)
print("ETAPA 9: Verificação Final...")
print("=" * 80)

# Verificar arquivo gerado
if arquivo_saida.exists():
    tamanho_mb = arquivo_saida.stat().st_size / (1024 * 1024)
    print("✓ Arquivo gerado com sucesso!")
    print(f"  Caminho: {arquivo_saida}")
    print(f"  Tamanho: {tamanho_mb:.2f} MB")

    # Ler e validar
    df_verificacao = pd.read_parquet(arquivo_saida)
    print(f"\n  Total de registros: {len(df_verificacao)}")
    print(f"  Colunas: {df_verificacao.columns.tolist()}")
    print(f"  Tipos de dados:\n{df_verificacao.dtypes}")
    print("\n  Distribuição por id_tabela:")
    print(df_verificacao["id_tabela"].value_counts())

    print("\n  Amostra de dados:")
    print(df_verificacao.head(10).to_string())
else:
    print("✗ Erro: Arquivo não foi criado!")
    sys.exit(1)

# ===== RESUMO FINAL =====
print("\n" + "=" * 80)
print("✓ PROCESSAMENTO CONCLUÍDO COM SUCESSO!")
print("=" * 80)
print("\nResumo:")
print(f"  - Arquivo Parquet: {arquivo_saida}")
print(f"  - Total de registros: {len(df_final)}")
print(f"  - Registros V4010: {len(v4010_final)}")
print(
    f"  - Registros de educacao: {(df_final['id_tabela'] == 'educacao').sum()}"
)
print(
    f"  - Registros de microdados: {(df_final['id_tabela'] == 'microdados').sum()}"
)
print("  - Cobertura temporal preenchida: 100% (todos com '(1)')")
if len(v4010_final) > 0:
    v4010_com_zeros_count = len(
        v4010_final[v4010_final["chave"].astype(str).str.startswith("0")]
    )
    print(
        f"  - Zeros à esquerda preservados: {v4010_com_zeros_count} registros"
    )
print("\n" + "=" * 80)
