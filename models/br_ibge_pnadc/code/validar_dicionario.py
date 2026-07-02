"""
Script de Validação do Dicionário PNADC em Parquet
===================================================

Este script valida se todos os dados dos arquivos Excel foram corretamente
processados e incluídos no arquivo Parquet final.

Comparações realizadas:
1. Total de registros do PNADC vs Parquet
2. Total de registros de ocupação (V4010) vs Parquet
3. Variáveis únicas em cada fonte
4. Categorias por variável
5. Integridade de dados

Autor: Data Engineering Team
Data: 2026-05-13
"""

import platform
import sys
from pathlib import Path

# pyrefly: ignore [untyped-import]
import openpyxl
import pandas as pd

# ===== CONFIGURAÇÕES =====
print("=" * 100)
print("VALIDAÇÃO DO DICIONÁRIO PNADC EM PARQUET")
print("=" * 100)

# Detectar sistema operacional e ajustar caminhos

is_windows = platform.system() == "Windows"

if is_windows:
    base_path = Path(r"D:\Devops\Git\BD\areadetrabalho")
    arquivo_pnadc = base_path / "dicionario_PNADC_microdados_trimestral.xls"
    arquivo_ocupacao = base_path / "Estrutura_Ocupacao_COD.xlsx"
    arquivo_parquet = base_path / "dicionario_dbt_final.parquet"
else:
    base_path = Path("/home/ubuntu/upload")
    arquivo_pnadc = base_path / "dicionario_PNADC_microdados_trimestral.xls"
    arquivo_ocupacao = base_path / "Estrutura_Ocupacao_COD.xlsx"
    arquivo_parquet = Path("/home/ubuntu/output/dicionario_dbt_final.parquet")

print(f"\nSistema Operacional: {platform.system()}")
print(f"Arquivo PNADC: {arquivo_pnadc}")
print(f"Arquivo Ocupação: {arquivo_ocupacao}")
print(f"Arquivo Parquet: {arquivo_parquet}")

# ===== ETAPA 1: Ler Parquet =====
print("\n" + "=" * 100)
print("ETAPA 1: Lendo arquivo Parquet...")
print("=" * 100)

try:
    df_parquet = pd.read_parquet(arquivo_parquet)
    print("✓ Arquivo Parquet lido com sucesso")
    print(f"  Total de registros: {len(df_parquet)}")
    print(f"  Colunas: {df_parquet.columns.tolist()}")
except Exception as e:
    print(f"✗ Erro ao ler Parquet: {e}")
    sys.exit(1)

# ===== ETAPA 2: Ler PNADC Original =====
print("\n" + "=" * 100)
print("ETAPA 2: Lendo arquivo PNADC original...")
print("=" * 100)

try:
    df_pnadc_raw = pd.read_excel(arquivo_pnadc, sheet_name=0, header=None)
    print("✓ Arquivo PNADC lido com sucesso")
    print(f"  Total de linhas: {len(df_pnadc_raw)}")
except Exception as e:
    print(f"✗ Erro ao ler PNADC: {e}")
    sys.exit(1)

# Processar PNADC original para comparação
registros_pnadc_original = []
id_tabela_atual = "microdados"

for _, row in df_pnadc_raw.iterrows():
    col0 = str(row[0]).strip() if pd.notna(row[0]) else ""
    col2 = str(row[2]).strip() if pd.notna(row[2]) else ""
    col5 = str(row[5]).strip() if pd.notna(row[5]) else ""
    col6 = str(row[6]).strip() if pd.notna(row[6]) else ""

    if "Parte" in col0:
        if "Características de educação" in col0:
            id_tabela_atual = "educacao"
        else:
            id_tabela_atual = "microdados"
        continue

    if col2 in [
        "Dicionário das variáveis da PNAD contínua",
        "Posição inicial",
        "Código da variável",
        "Tipo",
        "Descrição",
    ]:
        continue

    if not col2 or col2 == "NaN":
        continue

    if (
        col2
        and col6
        and col6.lower() not in ["não aplicável", "nao aplicavel"]
    ):
        registros_pnadc_original.append(
            {
                "id_tabela": id_tabela_atual,
                "nome_coluna": col2,
                "chave": col5,
                "valor": col6,
            }
        )

df_pnadc_original = pd.DataFrame(registros_pnadc_original)
print(f"  Total de registros processados: {len(df_pnadc_original)}")
print(
    f"  Registros de educacao: {(df_pnadc_original['id_tabela'] == 'educacao').sum()}"
)
print(
    f"  Registros de microdados: {(df_pnadc_original['id_tabela'] == 'microdados').sum()}"
)

# ===== ETAPA 3: Ler Ocupação Original =====
print("\n" + "=" * 100)
print("ETAPA 3: Lendo arquivo de Ocupação original...")
print("=" * 100)

try:
    wb = openpyxl.load_workbook(arquivo_ocupacao)
    ws = wb.active
    print("✓ Arquivo de Ocupação lido com sucesso")
    # pyrefly: ignore [missing-attribute]
    print(f"  Total de linhas: {ws.max_row}")
except Exception as e:
    print(f"✗ Erro ao ler Ocupação: {e}")
    sys.exit(1)

# Processar Ocupação original para comparação
registros_ocupacao_original = []
chaves_processadas = set()

# pyrefly: ignore [missing-attribute]
for row_idx in range(3, ws.max_row + 1):
    # pyrefly: ignore [unsupported-operation]
    gg = ws[f"A{row_idx}"].value
    # pyrefly: ignore [unsupported-operation]
    sg = ws[f"B{row_idx}"].value
    # pyrefly: ignore [unsupported-operation]
    sub = ws[f"C{row_idx}"].value
    # pyrefly: ignore [unsupported-operation]
    gb = ws[f"D{row_idx}"].value
    # pyrefly: ignore [unsupported-operation]
    den = ws[f"E{row_idx}"].value

    chave = None
    if gb is not None:
        chave = str(gb)
    elif sub is not None:
        chave = str(sub)
    elif sg is not None:
        chave = str(sg)
    elif gg is not None:
        chave = str(gg)

    if chave and den is not None:
        den_str = str(den).strip().lower()

        # Combinando as condições em um único bloco if
        if (
            den_str not in ["nao aplicavel", "não aplicável"]
            and "codigo" not in den_str
            and chave not in chaves_processadas
        ):
            chaves_processadas.add(chave)

            registros_ocupacao_original.append(
                {
                    "nome_coluna": "V4010",
                    "chave": chave,
                    "valor": str(den).strip(),
                }
            )

df_ocupacao_original = pd.DataFrame(registros_ocupacao_original)
print(f"  Total de registros V4010 processados: {len(df_ocupacao_original)}")
print(f"  Chaves únicas: {df_ocupacao_original['chave'].nunique()}")

# ===== ETAPA 4: Comparação de Totais =====
print("\n" + "=" * 100)
print("ETAPA 4: Comparação de Totais")
print("=" * 100)

# Remover V4010 do PNADC para comparação justa
df_pnadc_sem_v4010 = df_pnadc_original[
    df_pnadc_original["nome_coluna"] != "V4010"
].copy()

total_esperado = len(df_pnadc_sem_v4010) + len(df_ocupacao_original)
total_parquet = len(df_parquet)

print(f"\nRegistros PNADC (sem V4010): {len(df_pnadc_sem_v4010)}")
print(f"Registros V4010 (Ocupação): {len(df_ocupacao_original)}")
print(f"Total esperado: {total_esperado}")
print(f"Total no Parquet: {total_parquet}")

if total_esperado == total_parquet:
    print("✓ TOTAIS CONFEREM!")
else:
    diferenca = total_parquet - total_esperado
    print(f"⚠️ DIFERENÇA: {diferenca} registros")
    if diferenca > 0:
        print(f"   O Parquet tem {diferenca} registros a mais")
    else:
        print(f"   O Parquet tem {abs(diferenca)} registros a menos")

# ===== ETAPA 5: Comparação de Variáveis =====
print("\n" + "=" * 100)
print("ETAPA 5: Comparação de Variáveis Únicas")
print("=" * 100)

variaveis_pnadc = set(df_pnadc_original["nome_coluna"].unique())
variaveis_parquet = set(df_parquet["nome_coluna"].unique())

print(f"\nVariáveis no PNADC original: {len(variaveis_pnadc)}")
print(f"Variáveis no Parquet: {len(variaveis_parquet)}")

variaveis_faltando = variaveis_pnadc - variaveis_parquet
variaveis_extras = variaveis_parquet - variaveis_pnadc

if variaveis_faltando:
    print(f"\n⚠️ VARIÁVEIS FALTANDO no Parquet ({len(variaveis_faltando)}):")
    for var in sorted(variaveis_faltando):
        print(f"  - {var}")
else:
    print("\n✓ Todas as variáveis do PNADC estão no Parquet")

if variaveis_extras:
    print(f"\n⚠️ VARIÁVEIS EXTRAS no Parquet ({len(variaveis_extras)}):")
    for var in sorted(variaveis_extras):
        print(f"  - {var}")
else:
    print("\n✓ Nenhuma variável extra no Parquet")

# ===== ETAPA 6: Comparação de Categorias por Variável =====
print("\n" + "=" * 100)
print("ETAPA 6: Comparação de Categorias por Variável")
print("=" * 100)

# Verificar V4010 especificamente
print("\nVariável V4010:")
v4010_original = df_ocupacao_original["chave"].nunique()
v4010_parquet = len(df_parquet[df_parquet["nome_coluna"] == "V4010"])

print(f"  Categorias no original: {v4010_original}")
print(f"  Registros no Parquet: {v4010_parquet}")

if v4010_original == v4010_parquet:
    print("  ✓ V4010 CONFEREM!")
else:
    print(f"  ⚠️ DIFERENÇA: {v4010_parquet - v4010_original} registros")

# Verificar algumas variáveis importantes
variaveis_importantes = [
    "UF",
    "Capital",
    "RM_RIDE",
    "Trimestre",
    "Ano",
    "V3001",
    "V3002",
    "V3003",
]

print("\nComparação de categorias para variáveis importantes:")
for var in variaveis_importantes:
    if var in variaveis_pnadc:
        categorias_original = df_pnadc_original[
            df_pnadc_original["nome_coluna"] == var
        ]["chave"].nunique()
        categorias_parquet = len(df_parquet[df_parquet["nome_coluna"] == var])

        status = "✓" if categorias_original == categorias_parquet else "⚠️"
        print(
            f"  {status} {var}: {categorias_original} original vs {categorias_parquet} Parquet"
        )

# ===== ETAPA 7: Verificação de Integridade =====
print("\n" + "=" * 100)
print("ETAPA 7: Verificação de Integridade")
print("=" * 100)

# Verificar colunas obrigatórias
colunas_obrigatorias = [
    "id_tabela",
    "nome_coluna",
    "chave",
    "cobertura_temporal",
    "valor",
]
print("\nColunas obrigatórias:")
for col in colunas_obrigatorias:
    if col in df_parquet.columns:
        print(f"  ✓ {col}")
    else:
        print(f"  ✗ {col} - FALTANDO!")

# Verificar valores nulos
print("\nValores nulos por coluna:")
for col in df_parquet.columns:
    nulos = df_parquet[col].isna().sum()
    if nulos > 0:
        print(f"  ⚠️ {col}: {nulos} valores nulos")
    else:
        print(f"  ✓ {col}: sem valores nulos")

# Verificar cobertura_temporal
print("\nVerificação de cobertura_temporal:")
valores_cobertura = df_parquet["cobertura_temporal"].unique()
print(f"  Valores únicos: {valores_cobertura}")
if len(valores_cobertura) == 1 and valores_cobertura[0] == "(1)":
    print("  ✓ Todos os registros têm cobertura_temporal = '(1)'")
else:
    print("  ⚠️ Valores diferentes encontrados")

# Verificar zeros à esquerda em V4010
print("\nVerificação de zeros à esquerda em V4010:")
v4010_parquet_data = df_parquet[df_parquet["nome_coluna"] == "V4010"]
v4010_com_zeros = v4010_parquet_data[
    v4010_parquet_data["chave"].astype(str).str.startswith("0")
]
print(f"  Total de V4010: {len(v4010_parquet_data)}")
print(f"  Com zeros à esquerda: {len(v4010_com_zeros)}")

zeros_unicos = sorted(
    [z for z in v4010_parquet_data["chave"].unique() if str(z).startswith("0")]
)
if len(zeros_unicos) > 0:
    print(f"  Valores com zeros: {zeros_unicos}")
    print("  ✓ Zeros à esquerda preservados")
else:
    print("  ⚠️ Nenhum valor com zeros encontrado")

# ===== ETAPA 8: Distribuição por id_tabela =====
print("\n" + "=" * 100)
print("ETAPA 8: Distribuição por id_tabela")
print("=" * 100)

print("\nDistribuição no Parquet:")
print(df_parquet["id_tabela"].value_counts())

print("\nDistribuição esperada:")
print(f"  educacao: {(df_pnadc_original['id_tabela'] == 'educacao').sum()}")
print(
    f"  microdados: {(df_pnadc_original['id_tabela'] == 'microdados').sum() + len(df_ocupacao_original)}"
)

# ===== ETAPA 9: Amostra de Dados =====
print("\n" + "=" * 100)
print("ETAPA 9: Amostra de Dados do Parquet")
print("=" * 100)

print("\nPrimeiros 10 registros:")
print(df_parquet.head(10).to_string())

print("\n\nÚltimos 10 registros:")
print(df_parquet.tail(10).to_string())

print("\n\nAmostra de V4010 com zeros:")
print(v4010_com_zeros.head(10).to_string())

# ===== RESUMO FINAL =====
print("\n" + "=" * 100)
print("RESUMO FINAL DE VALIDAÇÃO")
print("=" * 100)

validacoes = {
    "Totais conferem": total_esperado == total_parquet,
    "Variáveis PNADC presentes": len(variaveis_faltando) == 0,
    "Sem variáveis extras": len(variaveis_extras) == 0,
    "V4010 completo": v4010_original == v4010_parquet,
    "Colunas obrigatórias": all(
        col in df_parquet.columns for col in colunas_obrigatorias
    ),
    "Sem valores nulos": df_parquet.isna().sum().sum() == 0,
    "Cobertura temporal preenchida": len(valores_cobertura) == 1
    and valores_cobertura[0] == "(1)",
    "Zeros à esquerda preservados": len(v4010_com_zeros) > 0,
}

print("\nResultados:")
for validacao, resultado in validacoes.items():
    status = "✓" if resultado else "✗"
    print(f"  {status} {validacao}")

todas_validas = all(validacoes.values())

print("\n" + "=" * 100)
if todas_validas:
    print("✓ VALIDAÇÃO COMPLETA - TODOS OS DADOS ESTÃO CORRETOS!")
else:
    print("⚠️ VALIDAÇÃO COM PROBLEMAS - VERIFIQUE OS ITENS ACIMA")
print("=" * 100)
