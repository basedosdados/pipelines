"""Build the per-table auxiliary-file bundles for us_epa_ghgrp.

One ZIP per data table under ``$EPA_GHGRP_DATA_DIR/auxiliary_files/<table>/``,
holding what a user of *that* table needs to read it: EPA's own FAQ on the
published GHGRP data (facility ids, FRS ids, biogenic CO2, sector vs subpart
totals, confidential values, GWPs), the subpart / industry-type list, and a
README with provenance, download dates and the transformations applied on load.

Both documents are lifted from the "FAQs about this Data" and "Industry Type"
sheets of EPA's Data Summary Spreadsheets workbook (``ghgp_data_<year>.xlsx``),
which must already be unzipped under ``<data_dir>/input/summary/``.

The long-form references — the Envirofacts GHG data model, the 40 CFR Part 98
subpart resource pages, the GWP table — are link-only: stable at EPA and read
once. They are indexed in each README.

Usage:
    uv run python models/us_epa_ghgrp/code/build_auxiliary_files.py
    gcloud storage cp -r <data_dir>/auxiliary_files/* gs://basedosdados/auxiliary_files/us_epa_ghgrp/
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import os
import zipfile
from pathlib import Path

import openpyxl

DATASET = "us_epa_ghgrp"
DATA_DIR = Path(
    os.environ.get(
        "EPA_GHGRP_DATA_DIR", Path.home() / "Downloads" / "us_epa_ghgrp_data"
    )
)
AUX = DATA_DIR / "auxiliary_files"
SUMMARY_ZIP_URL = (
    "https://www.epa.gov/system/files/other-files/2024-10/"
    "2023_data_summary_spreadsheets.zip"
)

LINKS = [
    (
        "GHGRP data sets (FLIGHT summary spreadsheets and subpart files)",
        "https://www.epa.gov/ghgreporting/data-sets",
    ),
    (
        "Envirofacts GHG data model (the pub_* tables this dataset is built from)",
        "https://www.epa.gov/enviro/greenhouse-gas-model",
    ),
    (
        "Envirofacts REST API documentation",
        "https://www.epa.gov/enviro/envirofacts-data-service-api",
    ),
    (
        "Resources by subpart (40 CFR Part 98)",
        "https://www.epa.gov/ghgreporting/resources-subpart-ghg-reporting",
    ),
    (
        "Global warming potentials used by the GHGRP (Table A-1 of 40 CFR 98)",
        "https://www.epa.gov/ghgreporting/ghgrp-global-warming-potentials",
    ),
    (
        "FLIGHT (Facility Level Information on GreenHouse gases Tool)",
        "https://ghgdata.epa.gov/ghgp/main.do",
    ),
]

TABLE_NOTES = {
    "facility": (
        "One row per GHGRP facility and reporting year, from `pub_dim_facility`. "
        "Kept as published, except: `state_id` is derived from the reported state "
        "abbreviation (the API carries no state FIPS); four-character ZIP codes are "
        "left-padded to five; `address` joins address1 and address2. Facilities that "
        "stopped reporting are carried by EPA with `reporting_status` set and no "
        "submission — they have no emission rows for that year. About 0.5% of county "
        "FIPS codes sit in a different state from the reported state (corporate-office "
        "counties on basin-level reporters); both are kept as reported."
    ),
    "emission_subpart": (
        "Facility x year x subpart x gas, from `pub_facts_subp_ghg_emission`, with the "
        "dimension ids replaced by their codes (subpart letter, gas code). Values are "
        "metric tons of CO2 equivalent. Biogenic CO2 (`gas = BIOCO2`) is published "
        "separately and is not part of EPA's facility totals. Null emissions are "
        "confidential business information withheld by EPA (subpart UU)."
    ),
    "emission_sector": (
        "Facility x year x sector x subsector x gas, from "
        "`pub_facts_sector_ghg_emission`, with the dimension ids replaced by their "
        "codes. Sectors are the FLIGHT dashboard classification; stationary combustion "
        "(subpart C) is attributed to the facility's sector, so sector totals differ "
        "from subpart totals (see FAQ 5). For the ~80 keys the API publishes as two "
        "rows, the rows were summed — that reproduces the subpart-table facility totals "
        "exactly. Rows with neither a gas nor a value (placeholders for facilities that "
        "did not report that year) were dropped; the remaining null emissions are "
        "confidential values withheld by EPA."
    ),
}


def read_sheets(xlsx: Path) -> tuple[str, list[list[str]]]:
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    faq = [
        str(row[0]).strip()
        for row in wb["FAQs about this Data"].iter_rows(values_only=True)
        if row[0]
    ]
    industry = [
        [str(c).strip() for c in row[2:5]]
        for row in wb["Industry Type"].iter_rows(values_only=True)
        if row[2]
    ]
    return "\n\n".join(faq), industry


def readme(table: str, workbook: str, today: str) -> str:
    links = "\n".join(f"- {title}: {url}" for title, url in LINKS)
    return f"""# {DATASET} / {table} — auxiliary files

Source: U.S. Environmental Protection Agency, Greenhouse Gas Reporting Program
(GHGRP). Data read from the Envirofacts GHG REST API
(https://data.epa.gov/efservice/) on {today}. Public domain (work of the U.S.
federal government).

Suggested citation: U.S. EPA, Greenhouse Gas Reporting Program (GHGRP),
https://www.epa.gov/ghgreporting, reporting years 2010 onward, accessed {today}.

## Files in this bundle

- `ghgrp_data_faq.md` — EPA's "FAQs about this Data", copied verbatim from the
  sheet of the same name in `{workbook}` (Data Summary Spreadsheets,
  {SUMMARY_ZIP_URL}), downloaded {today}. Explains facility and FRS ids, NAICS
  codes, biogenic CO2, sector vs subpart totals, CEMS, direct emitters vs
  suppliers, confidential values, CO2 injection, basin-level addresses and the
  global warming potentials used.
- `subpart_industry_types.csv` — the "Industry Type" sheet of the same workbook:
  subpart letter, industry name and facility type for every reporting category,
  including the industry-type suffixes (W-ONSH, MM-REF, NN-LDC, ...) that appear
  in `facility.industry_type`.

## How this table was built

{TABLE_NOTES[table]}

## Link-only references

{links}
"""


def main() -> None:
    summary_dir = DATA_DIR / "input" / "summary"
    workbooks = sorted(summary_dir.glob("ghgp_data_20??.xlsx"))
    if not workbooks:
        raise FileNotFoundError(
            f"no ghgp_data_<year>.xlsx under {summary_dir}"
        )
    xlsx = workbooks[-1]
    faq, industry = read_sheets(xlsx)
    today = dt.date.today().isoformat()

    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(["subpart", "industry_name", "facility_type"])
    writer.writerows(industry[1:])  # first row is the sheet header
    industry_csv = buf.getvalue()

    for table in TABLE_NOTES:
        dest = AUX / table / "auxiliary_files.zip"
        dest.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("README.md", readme(table, xlsx.name, today))
            zf.writestr(
                "ghgrp_data_faq.md",
                "# FAQs about this data (EPA)\n\n" + faq + "\n",
            )
            zf.writestr("subpart_industry_types.csv", industry_csv)
        print(f"{table:18s} -> {dest} ({dest.stat().st_size:,} bytes)")
    print(
        f"\nUpload with:\n  gcloud storage cp -r {AUX}/* "
        f"gs://basedosdados/auxiliary_files/{DATASET}/"
    )


if __name__ == "__main__":
    main()
