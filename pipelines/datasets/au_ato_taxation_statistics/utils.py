"""Pure download and cleaning helpers for au_ato_taxation_statistics.

No Prefect imports here: this module is shared verbatim between the
one-shot onboarding bootstrap (``models/au_ato_taxation_statistics/code``)
and the recurring Prefect flow.

The ATO publishes one CKAN package per financial year, each holding ~96
Excel workbooks. Every detailed table has the same shape: a few leading
dimension columns followed by measure columns that come in
``<item> no.`` / ``<item> $`` pairs. The transform melts those pairs into
a long ``item`` / ``record_count`` / ``amount`` triple.
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any, cast

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import requests

from pipelines.datasets.au_ato_taxation_statistics.constants import constants

# Header cell ends with the measure suffix, separated by a newline or a
# space, optionally followed by a footnote digit: "Total income\nno.",
# "Net rent - profit no.", "Gross GST payable|$".
MEASURE_SUFFIX = re.compile(r"[\n ]\s*(no\.|\$)\s*\d*\s*$")
# "Number of individuals" is the bare-count spelling used before 2018-19.
BARE_COUNT = re.compile(r"^number of (individuals|companies)\s*\d*$", re.I)
# Footnote marker: digits glued directly onto a letter or a closing
# bracket ("Net rent4", "Statistical Area Level 4 (SA4)2"). Anchored so
# that "Other income category 1", "Subtotal 2" and "Tax loss 2022-23
# carried back to 2021-22" survive intact.
FOOTNOTE = re.compile(r"(?<=[a-zA-Z)])\d{1,2}$")
# Trailing "..., 2024-25 financial year" / "..., 2023-24 income year".
TITLE_YEAR = re.compile(
    r"(20\d{2})\s*[–—-]\s*(\d{2})\s*(?:income|financial)\s*year",  # noqa: RUF001
    re.I,
)
# Leading sort/classification prefix: "A. Mining", "011 Nursery", "ab. $6,001 ...".
VALUE_PREFIX = re.compile(r"^\s*([A-Za-z]{1,3}|\d{3,5})\.?\s+(.*)$")


def _clean_label(value: Any) -> str:
    """Normalise a dimension label: whitespace, footnote digit, stray case.

    Purely numeric values (postcodes, industry codes) are returned as-is so
    their digits are never mistaken for a footnote marker.
    """
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text or text.isdigit():
        return text
    text = re.sub(r"\bTO\b", "to", text)
    return FOOTNOTE.sub("", text).strip()


def _norm(text: Any) -> str:
    """Collapse whitespace and strip a trailing footnote digit.

    Args:
        text: Raw cell or header value of any type.

    Returns:
        The normalised text.
    """
    text = str(text).replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return FOOTNOTE.sub("", text).strip()


def clean_item(header: Any) -> str:
    """Turn a measure header into its canonical item label."""
    base = MEASURE_SUFFIX.sub("", str(header))
    base = _norm(base)
    if BARE_COUNT.match(base):
        return "Companies" if "compan" in base.lower() else "Individuals"
    return constants.ITEM_ALIASES.value.get(base.lower(), base)


def measure_kind(header: Any) -> str | None:
    """Return ``"count"``, ``"amount"`` or ``None`` for a header cell."""
    header = str(header)
    match = MEASURE_SUFFIX.search(header)
    if match:
        return "count" if match.group(1) == "no." else "amount"
    if BARE_COUNT.match(_norm(header)):
        return "count"
    return None


def split_prefix(value: Any) -> tuple[str | None, str | None]:
    """Split ``"A. Mining"`` into ``("A", "Mining")``.

    Returns ``(None, label)`` when the value carries no prefix.
    """
    if value is None:
        return None, None
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return None, None
    # "A.Mining" (no space after the dot) occurs in some releases.
    tight = re.match(r"^([A-Za-z]{1,3})\.(\S.*)$", text)
    if tight:
        return tight.group(1).upper(), tight.group(2).strip()
    match = VALUE_PREFIX.match(text)
    if match:
        code, label = match.group(1), match.group(2).strip()
        return (code.upper() if code.isalpha() else code), label
    return None, text


def data_year(title: Any) -> int | None:
    """Start year of the financial year the sheet actually covers.

    The release year is *not* reliable: the GST by-industry table shipped
    in the 2023-24 release covers the 2024-25 financial year.
    """
    matches = TITLE_YEAR.findall(str(title or ""))
    if not matches:
        return None
    return int(matches[-1][0])


def pick_sheet(workbook, table: str) -> str:
    """Choose the data worksheet for ``table`` within ``workbook``."""
    pattern = re.compile(constants.SHEET_SELECTORS.value[table], re.I)
    names = [s for s in workbook.sheetnames if pattern.search(s)]
    if not names:
        raise ValueError(
            f"no worksheet matching {table!r}: {workbook.sheetnames}"
        )
    return names[0]


def find_header_row(rows: list[tuple], scan: int = 6) -> int:
    """Index of the header row: the one with the most measure suffixes."""
    best, best_score = 0, -1
    for idx in range(min(scan, len(rows))):
        score = sum(
            1
            for cell in rows[idx]
            if cell is not None and measure_kind(str(cell)) is not None
        )
        if score > best_score:
            best, best_score = idx, score
    if best_score <= 0:
        raise ValueError("no header row with measure columns found")
    return best


def _to_number(value: Any) -> float | None:
    """Coerce a cell to a number, treating ATO suppression markers as null."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if text in {"", "-", "–", "—", "n.a.", "n/a", "np", "NP", "*"}:  # noqa: RUF001
        return None
    try:
        return float(text)
    except ValueError:
        return None


def read_table(path: Path, table: str) -> pd.DataFrame:
    """Read one workbook and return the long-format frame for ``table``."""
    # pyrefly: ignore [untyped-import]
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = pick_sheet(workbook, table)
        worksheet = workbook[sheet]
        rows = list(worksheet.iter_rows(values_only=True))
        title = next((c for c in rows[0] if c), "") if rows else ""
        header_idx = find_header_row(rows)
        header = ["" if c is None else str(c) for c in rows[header_idx]]
    finally:
        workbook.close()

    year = data_year(title)
    if year is None:
        raise ValueError(f"{path.name}: cannot parse data year from {title!r}")

    dim_cols: list[tuple[int, str]] = []
    meas_cols: list[tuple[int, str, str]] = []
    for idx, cell in enumerate(header):
        if not cell.strip():
            continue
        kind = measure_kind(cell)
        if kind:
            meas_cols.append((idx, clean_item(cell), kind))
        else:
            key = _norm(cell).lower()
            name = constants.DIMENSION_NAMES.value.get(key)
            if name:
                dim_cols.append((idx, name))

    if not meas_cols:
        raise ValueError(f"{path.name}: no measure columns detected")

    # Group the no./$ pair belonging to the same item, preserving order.
    items: dict[str, dict[str, int]] = {}
    for idx, item, kind in meas_cols:
        items.setdefault(item, {})[kind] = idx

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if row is None or all(c is None for c in row):
            continue
        dims: dict[str, Any] = {}
        for idx, name in dim_cols:
            value = row[idx] if idx < len(row) else None
            suffix = constants.PREFIXED_DIMENSIONS.value.get(name)
            if suffix:
                code, label = split_prefix(value)
                dims[f"{name}_{suffix}"] = code
                dims[name] = _clean_label(label) if label else None
            elif value is None:
                dims[name] = None
            else:
                dims[name] = _clean_label(value) or None
        # A row whose dimensions are all empty is a spacer or a footnote.
        if not any(v for v in dims.values()):
            continue
        for item, slots in items.items():
            count = (
                _to_number(row[slots["count"]])
                if "count" in slots and slots["count"] < len(row)
                else None
            )
            amount = (
                _to_number(row[slots["amount"]])
                if "amount" in slots and slots["amount"] < len(row)
                else None
            )
            if count is None and amount is None:
                continue
            records.append(
                {
                    "year": year,
                    **dims,
                    "item": item,
                    "record_count": None if count is None else round(count),
                    "amount": amount,
                }
            )

    frame = pd.DataFrame.from_records(records)
    if frame.empty:
        return frame
    columns = [
        "year",
        *constants.DIMENSIONS.value[table],
        *constants.MEASURES.value,
    ]
    for column in columns:
        if column not in frame.columns:
            frame[column] = None
    return frame[columns]


def build_dicionario(frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Value/label dictionary for the coded dimension columns."""
    rows: list[dict[str, Any]] = []
    for table, frame in frames.items():
        if frame.empty:
            continue
        for column in constants.DICTIONARY_COLUMNS.value.get(table, []):
            if column not in frame.columns:
                continue
            label_col = None
            for suffix in ("_id", "_code"):
                if column.endswith(suffix):
                    label_col = column[: -len(suffix)]
                    break
            if label_col and label_col in frame.columns:
                pairs = (
                    frame[[column, label_col]]
                    .dropna(subset=[column])
                    .drop_duplicates()
                    .itertuples(index=False)
                )
                seen: dict[str, str] = {}
                for key, label in pairs:
                    seen.setdefault(str(key), str(label))
                items = sorted(seen.items())
            else:
                items = [
                    (str(v), str(v))
                    for v in sorted(frame[column].dropna().unique())
                ]
            years = sorted(frame["year"].dropna().unique())
            coverage = (
                f"{int(min(years))}(1){int(max(years))}" if years else ""
            )
            for key, label in items:
                rows.append(
                    {
                        "id_tabela": table,
                        "nome_coluna": column,
                        "chave": key,
                        "cobertura_temporal": coverage,
                        "valor": label,
                    }
                )
    return pd.DataFrame(
        rows,
        columns=[
            "id_tabela",
            "nome_coluna",
            "chave",
            "cobertura_temporal",
            "valor",
        ],
    )


def to_string_table(frame: pd.DataFrame, columns: list[str]) -> pa.Table:
    """Cast a frame to an all-STRING arrow table with a stable column order.

    Staging is all-STRING by house convention and the dbt model
    ``safe_cast``s each column. The cast goes through arrow rather than
    ``astype(str)``, which would render NULL as the literal ``"nan"``.
    Integer columns are passed through int64 first so ``year`` serialises
    as ``"2023"`` and not ``"2023.0"``.
    """
    arrays = []
    for column in columns:
        series = (
            frame[column]
            if column in frame.columns
            else pd.Series([None] * len(frame))
        )
        if column in {"year", "record_count"}:
            array = pa.array(series, type=pa.int64(), from_pandas=True)
        elif column == "amount":
            array = pa.array(series, type=pa.float64(), from_pandas=True)
        else:
            values = [None if pd.isna(v) else str(v) for v in series]
            array = pa.array(values, type=pa.string())
        arrays.append(array.cast(pa.string()))
    return pa.Table.from_arrays(arrays, names=columns)


def write_partitioned(
    frame: pd.DataFrame, columns: list[str], out_dir: Path
) -> int:
    """Write ``frame`` as ``year=<Y>/data.parquet`` under ``out_dir``."""
    written = 0
    for year_key, part in frame.groupby("year", sort=True):
        year = cast(int, year_key)
        target = out_dir / f"year={year}"
        target.mkdir(parents=True, exist_ok=True)
        table = to_string_table(
            part.drop(columns=["year"]), [c for c in columns if c != "year"]
        )
        pq.write_table(table, target / "data.parquet", compression="snappy")
        written += part.shape[0]
    return written


def list_packages(session: requests.Session | None = None) -> list[dict]:
    """All ATO CKAN packages named ``taxation-statistics-<YYYY-YY>``."""
    session = session or requests.Session()
    packages: list[dict] = []
    start = 0
    while True:
        response = session.get(
            constants.CKAN_API.value,
            params={
                "fq": f"organization:{constants.CKAN_ORG.value}",
                "rows": 100,
                "start": start,
            },
            headers={"User-Agent": constants.USER_AGENT.value},
            timeout=120,
        )
        response.raise_for_status()
        result = response.json()["result"]
        packages.extend(result["results"])
        start += 100
        if start >= result["count"]:
            break
    return [
        p
        for p in packages
        if re.fullmatch(r"taxation-statistics-20\d{2}-\d{2}", p["name"])
    ]


def resource_map(packages: list[dict]) -> dict[str, dict[str, str]]:
    """Map ``{release: {table: url}}`` for the curated table subset."""
    mapping: dict[str, dict[str, str]] = {}
    for package in packages:
        release = package["name"].replace("taxation-statistics-", "")
        for table, pattern in constants.TABLE_SELECTORS.value.items():
            regex = re.compile(pattern, re.I)
            hits = [
                r["url"]
                for r in package["resources"]
                if str(r.get("url", "")).lower().endswith(".xlsx")
                and regex.search(str(r.get("url", "")).rsplit("/", 1)[-1])
            ]
            if len(hits) == 1:
                mapping.setdefault(release, {})[table] = hits[0]
    return mapping


def download(
    url: str, dest: Path, session: requests.Session | None = None
) -> Path:
    """Download ``url`` to ``dest`` (skipped when already present)."""
    if dest.exists() and dest.stat().st_size > 5000:
        return dest
    session = session or requests.Session()
    dest.parent.mkdir(parents=True, exist_ok=True)
    response = session.get(
        url, headers={"User-Agent": constants.USER_AGENT.value}, timeout=600
    )
    response.raise_for_status()
    dest.write_bytes(response.content)
    return dest


def clean_all(input_dir: Path, output_dir: Path) -> dict[str, int]:
    """Clean every downloaded workbook into partitioned parquet.

    Input files are named ``<table>__<release>.xlsx``.
    """
    counts: dict[str, int] = {}
    frames: dict[str, pd.DataFrame] = {}
    for table in constants.TABLE_SELECTORS.value:
        parts: list[pd.DataFrame] = []
        for path in sorted(input_dir.glob(f"{table}__*.xlsx")):
            part = read_table(path, table)
            if not part.empty:
                parts.append(part)
        if not parts:
            raise FileNotFoundError(
                f"no input workbooks for {table!r} in {input_dir}; refusing to "
                "write a partial snapshot, since upload.py publishes the whole "
                "output tree and stale partitions would survive"
            )
        frame = pd.concat(parts, ignore_index=True)
        frames[table] = frame
        columns = [
            "year",
            *constants.DIMENSIONS.value[table],
            *constants.MEASURES.value,
        ]
        counts[table] = write_partitioned(frame, columns, output_dir / table)

    dicionario = build_dicionario(frames)
    if not dicionario.empty:
        target = output_dir / "dicionario"
        target.mkdir(parents=True, exist_ok=True)
        table = to_string_table(dicionario, list(dicionario.columns))
        pq.write_table(table, target / "data.parquet", compression="snappy")
        counts["dicionario"] = dicionario.shape[0]
    return counts


def iter_tables() -> Iterator[str]:
    """Names of every table produced by this dataset, dicionario last."""
    yield from constants.TABLE_SELECTORS.value
    yield "dicionario"
