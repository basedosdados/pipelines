"""Download + cleaning transform for au_abs_labour_force (shared by the pipeline
and the one-shot bootstrap in models/au_abs_labour_force/code/).

Pure functions (no Prefect) so they are importable and unit-testable. Schema and
column order come from the architecture CSVs (the single source of truth).

Sources
-------
- SDMX-CSV (ABS Data API), ``labels=both`` so each cell is ``"code: label"``:
    * ``labour_force_status`` <- ``LF`` (states, age total) + ``LF_AGES``
      (national, all ages, adds Not-in-labour-force + Civilian population).
    * ``underutilisation``    <- ``LF_UNDER`` (state and age).
- ABS time-series Excel spreadsheets (curated API does not serve these):
    * ``hours_worked``          <- Table 18 (national, by sex).
    * ``status_in_employment``  <- Table 19 (national) + SEM1 (states pivot).

ABS reports counts in thousands and hours in thousands of hours; every value is
scaled by ``10 ** UNIT_MULT`` (SDMX) or by the Index unit (Excel) to absolute
persons / hours, so the ``person`` / ``hour`` measurement units are truthful.
Rates are left untouched. National totals come from the national sources; states
from the state sources — ABS benchmarks national separately, so national is never
derived by summing states.
"""

import csv
import logging
from pathlib import Path

# pyrefly: ignore [untyped-import]
import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import requests

from pipelines.datasets.au_abs_labour_force.constants import constants

log = logging.getLogger("au_abs_labour_force")

PA = {"STRING": pa.string(), "INT64": pa.int64(), "FLOAT64": pa.float64()}
_ARCH = constants.ARCHITECTURE_DIR.value

# ── label normalisation ──────────────────────────────────────────────────────
SEX = {"males": "males", "females": "females", "persons": "persons"}
TSEST = {
    "original": "original",
    "seasonally adjusted": "seasonally_adjusted",
    "trend": "trend",
}

# SDMX MEASURE / PARM_ITEM code -> architecture column.
LFS_MEASURE = {
    "M3": "employed_total",
    "M1": "employed_full_time",
    "M2": "employed_part_time",
    "M6": "unemployed_total",
    "M4": "unemployed_looked_for_full_time",
    "M5": "unemployed_looked_for_part_time",
    "M9": "labour_force_total",
    "M10": "not_in_labour_force",
    "M11": "civilian_population_15_over",
    "M13": "unemployment_rate",
    "M14": "unemployment_rate_looked_for_full_time",
    "M15": "unemployment_rate_looked_for_part_time",
    "M12": "participation_rate",
    "M16": "employment_to_population_ratio",
}
UNDER_MEASURE = {
    "M21": "underemployed_total",
    "M22": "underemployment_ratio",
    "M23": "underemployment_rate",
    "M24": "underutilisation_rate",
}

# Status in employment (Excel labels) -> normalised label.
STATUS = {
    "employee": "employee",
    "owner manager of incorporated enterprise with employees": "owner manager of incorporated enterprise with employees",
    "owner manager of incorporated enterprise without employees": "owner manager of incorporated enterprise without employees",
    "owner manager of unincorporated enterprise with employees": "owner manager of unincorporated enterprise with employees",
    "owner manager of unincorporated enterprise without employees": "owner manager of unincorporated enterprise without employees",
    "contributing family worker": "contributing family worker",
    "contributing family worker or payment in kind": "contributing family worker",
}


def _norm_age(label: str) -> str:
    """Normalise an ABS age label, e.g. '15 - 24 years' -> '15-24 years'."""
    label = label.strip()
    if label.lower().startswith("total"):
        return "total"
    return label.replace(" - ", "-")


def _norm_hours_band(label: str) -> str:
    """Strip the ABS '>' / '>>' hierarchy markers from an hours-band label."""
    return label.lstrip("> ").strip()


def _year_month(period: str) -> tuple[int, int]:
    """Split an SDMX 'YYYY-MM' (or a pandas Timestamp) into (year, month)."""
    # pyrefly: ignore [unnecessary-type-conversion]
    s = str(period)[:7]
    y, m = s.split("-")
    return int(y), int(m)


# ── schema ───────────────────────────────────────────────────────────────────
def read_arch(table: str) -> list[dict]:
    """Read a table's architecture CSV — the schema source of truth."""
    with open(_ARCH / f"{table}.csv", newline="") as fh:
        return list(csv.DictReader(fh))


# ── SDMX download + parse ─────────────────────────────────────────────────────
def download_sdmx(flow: str, dest: Path) -> Path:
    """Download a full-history SDMX-CSV extract for one dataflow.

    ``/all`` returns every series and period, so the query is month-agnostic and
    the pipeline re-pulls the complete history each run (dump_mode overwrite).

    Args:
        flow: Dataflow id (``LF``, ``LF_AGES``, ``LF_UNDER``).
        dest: Directory to write ``<flow>.csv`` into; created if absent.

    Returns:
        Path to the written CSV.
    """
    dest.mkdir(parents=True, exist_ok=True)
    url = f"{constants.API_BASE.value}/{flow}/all"
    r = requests.get(
        url,
        headers={
            "User-Agent": constants.USER_AGENT.value,
            "Accept": constants.SDMX_ACCEPT.value,
        },
        timeout=600,
    )
    r.raise_for_status()
    out = dest / f"{flow}.csv"
    out.write_bytes(r.content)
    log.info(f"{flow}: {len(r.content):,} bytes -> {out}")
    return out


def _split_code(series: pd.Series) -> pd.Series:
    """From a 'code: label' SDMX-CSV cell, return the code (before ': ')."""
    return series.str.split(": ", n=1).str[0].str.strip()


def _split_label(series: pd.Series) -> pd.Series:
    """From a 'code: label' SDMX-CSV cell, return the label (after ': ')."""
    return series.str.split(": ", n=1).str[1].str.strip()


def load_sdmx(path: Path) -> pd.DataFrame:
    """Read one SDMX-CSV extract into a tidy long frame.

    Splits every ``code: label`` cell, scales OBS_VALUE by ``10 ** UNIT_MULT`` so
    counts become absolute persons and rates are untouched, and normalises the
    dimension labels.

    Returns:
        Columns: measure_code, sex, age_group, adjustment_type, region,
        year, month, value.
    """
    df = pd.read_csv(path, dtype=str, na_filter=False)
    # the measure dimension is MEASURE for LF/LF_AGES, PARM_ITEM for LF_UNDER.
    mcol = next(
        c for c in df.columns if c.split(":")[0] in ("MEASURE", "PARM_ITEM")
    )

    def col(prefix):
        return next(c for c in df.columns if c.split(":")[0] == prefix)

    out = pd.DataFrame(
        {
            "measure_code": _split_code(df[mcol]),
            "sex": _split_label(df[col("SEX")]).str.lower().map(SEX),
            "age_group": _split_label(df[col("AGE")]).map(_norm_age),
            "adjustment_type": _split_label(df[col("TSEST")])
            .str.lower()
            .map(TSEST),
            "region": _split_label(df[col("REGION")]),
        }
    )
    ym = df[col("TIME_PERIOD")].map(_year_month)
    out["year"] = ym.str[0]
    out["month"] = ym.str[1]
    mult = _split_code(df[col("UNIT_MULT")]).astype(int)
    out["value"] = pd.to_numeric(df["OBS_VALUE"], errors="coerce") * (
        10.0**mult
    )
    return out


def _pivot_measures(
    long: pd.DataFrame,
    measure_map: dict,
    keys: list[str],
    arch_cols: list[str],
) -> pd.DataFrame:
    """Pivot a long SDMX frame (measure_code -> value) to wide measure columns."""
    long = long[long["measure_code"].isin(measure_map)].copy()
    long["measure"] = long["measure_code"].map(measure_map)
    wide = long.pivot_table(
        index=keys, columns="measure", values="value", aggfunc="first"
    ).reset_index()
    wide.columns.name = None
    for c in arch_cols:
        if c not in wide.columns:
            wide[c] = pd.NA
    return wide[arch_cols]


def clean_labour_force_status(input_dir: Path) -> pd.DataFrame:
    """Build labour_force_status from LF + LF_AGES.

    LF carries the headline series for all nine regions at age total, in all
    three adjustments (Original/Seasonally Adjusted/Trend), but only 12 measures.
    LF_AGES adds the national age breakdown and the two measures LF omits
    (Not-in-labour-force, Civilian population) — but its total-age series is
    Original only. So both are needed: national SA/Trend at total age comes from
    LF, the age breakdown and M10/M11 from LF_AGES. The pivot merges the rows
    that overlap at (Australia, total age, Original), filling each measure once.
    """
    arch_cols = [a["name"] for a in read_arch("labour_force_status")]
    keys = [
        "year",
        "month",
        "geography",
        "sex",
        "age_group",
        "adjustment_type",
    ]

    lf = load_sdmx(input_dir / "LF.csv")  # all regions, age total, 3 adj
    lf["geography"] = lf["region"]

    national = load_sdmx(
        input_dir / "LF_AGES.csv"
    )  # national, all ages, +M10/M11
    national = national[national["region"] == "Australia"].copy()
    national["geography"] = "Australia"

    long = pd.concat([lf, national], ignore_index=True)
    return _pivot_measures(long, LFS_MEASURE, keys, arch_cols)


def clean_underutilisation(input_dir: Path) -> pd.DataFrame:
    """Build underutilisation from LF_UNDER (state and age, Orig/SA/Trend)."""
    arch_cols = [a["name"] for a in read_arch("underutilisation")]
    keys = [
        "year",
        "month",
        "geography",
        "sex",
        "age_group",
        "adjustment_type",
    ]
    long = load_sdmx(input_dir / "LF_UNDER.csv")
    long["geography"] = long["region"]
    return _pivot_measures(long, UNDER_MEASURE, keys, arch_cols)


# ── Excel time-series download + parse ────────────────────────────────────────
def download_excel(month: str, dest: Path) -> dict[str, Path]:
    """Download the Table 18 / Table 19 / SEM1 spreadsheets for a release month.

    Args:
        month: Release month slug, e.g. ``"jun-2026"``.
        dest: Directory to write into; created if absent.

    Returns:
        Mapping slug -> written path.
    """
    dest.mkdir(parents=True, exist_ok=True)
    base = constants.EXCEL_BASE.value.format(month=month)
    headers = {"User-Agent": constants.USER_AGENT.value}
    out = {}
    for slug, fname in constants.EXCEL_FILES.value.items():
        r = requests.get(f"{base}/{fname}", headers=headers, timeout=600)
        r.raise_for_status()
        p = dest / fname
        p.write_bytes(r.content)
        out[slug] = p
        log.info(f"{slug}: {len(r.content):,} bytes -> {p}")
    return out


def _index_meta(path: Path) -> pd.DataFrame:
    """Read an ABS time-series 'Index' sheet -> one row per series.

    Returns:
        Columns: series_id, description, unit, series_type.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Index"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(
        i
        for i, r in enumerate(rows)
        if r
        and any(isinstance(c, str) and c.strip() == "Series ID" for c in r)
    )
    header = [str(c).strip() if c is not None else "" for c in rows[hdr]]
    di = header.index("Data Item Description")
    si = header.index("Series ID")
    ui = header.index("Unit")
    ti = header.index("Series Type")
    recs = []
    for r in rows[hdr + 1 :]:
        if not r or si >= len(r) or r[si] is None:
            continue
        sid = str(r[si]).strip()
        if not sid or sid.startswith("©"):
            continue
        recs.append(
            {
                "series_id": sid,
                "description": str(r[di]).strip()
                if di < len(r) and r[di] is not None
                else "",
                "unit": str(r[ui]).strip()
                if ui < len(r) and r[ui] is not None
                else "",
                "series_type": str(r[ti]).strip()
                if ti < len(r) and r[ti] is not None
                else "",
            }
        )
    wb.close()
    return pd.DataFrame(recs)


def _timeseries_long(path: Path) -> pd.DataFrame:
    """Read all 'Data*' sheets of an ABS time-series spreadsheet -> long frame.

    Returns:
        Columns: series_id, year, month, value (value raw, unscaled).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    frames = []
    for sheet in [s for s in wb.sheetnames if s.startswith("Data")]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr = next(
            i
            for i, r in enumerate(rows)
            if r and isinstance(r[0], str) and r[0].strip() == "Series ID"
        )
        ids = [str(c).strip() if c is not None else "" for c in rows[hdr]]
        for r in rows[hdr + 1 :]:
            if not r or r[0] is None:
                continue
            # pyrefly: ignore [bad-argument-type]
            y, m = _year_month(pd.Timestamp(r[0]).strftime("%Y-%m"))
            for j in range(1, len(ids)):
                sid = ids[j]
                if not sid or j >= len(r) or r[j] is None:
                    continue
                frames.append((sid, y, m, r[j]))
    wb.close()
    df = pd.DataFrame(frames, columns=["series_id", "year", "month", "value"])
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    return df


def _apply_unit(df: pd.DataFrame) -> pd.DataFrame:
    """Scale values to absolute units from the Index 'unit': thousands -> x1000."""
    mult = (
        df["unit"]
        .str.strip()
        .str.lower()
        .map(lambda u: 1000.0 if u in ("000", "000 hours") else 1.0)
    )
    df = df.copy()
    df["value"] = df["value"] * mult
    return df


def clean_hours_worked(input_dir: Path) -> pd.DataFrame:
    """Build hours_worked from Table 18 (national, by sex and hours band)."""
    path = input_dir / constants.EXCEL_FILES.value["hours_worked"]
    arch_cols = [a["name"] for a in read_arch("hours_worked")]
    meta = _index_meta(path)
    data = _timeseries_long(path).merge(meta, on="series_id", how="left")
    data = _apply_unit(data)

    top = {
        "employed total": "employed_persons",
        "number of hours actually worked in all jobs": "hours_worked",
        "hours actually worked in all jobs per employed person": "hours_per_person",
    }
    recs = []
    for _, r in data.iterrows():
        segs = [s.strip() for s in r["description"].split(";") if s.strip()]
        if len(segs) < 2:
            continue
        if segs[0].lower() in top:  # band-less grand total
            band, measure, sex = "total", segs[0], segs[1]
        else:  # <band> ; <measure> ; <sex>
            if len(segs) < 3:
                continue
            band, measure, sex = segs[0], segs[1], segs[2]
        mcol = top.get(measure.lower())
        if mcol is None:
            continue
        recs.append(
            {
                "year": r["year"],
                "month": r["month"],
                "geography": "Australia",
                "sex": SEX.get(_norm_hours_band(sex).lower()),
                "hours_band": _norm_hours_band(band).lower()
                if band == "total"
                else _norm_hours_band(band),
                "measure": mcol,
                "value": r["value"],
            }
        )
    long = pd.DataFrame(recs)
    keys = ["year", "month", "geography", "sex", "hours_band"]
    wide = long.pivot_table(
        index=keys, columns="measure", values="value", aggfunc="first"
    ).reset_index()
    wide.columns.name = None
    for c in arch_cols:
        if c not in wide.columns:
            wide[c] = pd.NA
    return wide[arch_cols]


def clean_status_in_employment(input_dir: Path) -> pd.DataFrame:
    """Build status_in_employment from Table 19 (national) + SEM1 (states).

    National (persons, males, females) comes from Table 19; the eight states
    (males, females) come from SEM1 aggregated over hours bands, with persons
    derived as males + females within each state.
    """
    arch_cols = [a["name"] for a in read_arch("status_in_employment")]
    keys = ["year", "month", "geography", "sex", "status_in_employment"]

    # National — Table 19 time-series spreadsheet.
    path19 = input_dir / constants.EXCEL_FILES.value["status_national"]
    meta = _index_meta(path19)
    d19 = _apply_unit(
        _timeseries_long(path19).merge(meta, on="series_id", how="left")
    )
    emp_type = {
        "employed total": "employed_total",
        "employed full-time": "employed_full_time",
        "employed part-time": "employed_part_time",
    }
    recs = []
    for _, r in d19.iterrows():
        segs = [s.strip() for s in r["description"].split(";") if s.strip()]
        if len(segs) < 3:
            continue
        status = STATUS.get(segs[0].lower())
        etype = emp_type.get(_norm_hours_band(segs[1]).lower())
        sex = SEX.get(_norm_hours_band(segs[2]).lower())
        if status is None or etype is None or sex is None:
            continue
        recs.append(
            {
                "year": r["year"],
                "month": r["month"],
                "geography": "Australia",
                "sex": sex,
                "status_in_employment": status,
                "measure": etype,
                "value": r["value"],
            }
        )
    national = pd.DataFrame(recs)

    # States — SEM1 pivot (Data 1 sheet), aggregate over hours bands.
    sem = _read_sem1(input_dir / constants.EXCEL_FILES.value["status_states"])
    states = []
    for (y, m, geo, status), g in sem.groupby(
        ["year", "month", "geography", "status_in_employment"]
    ):
        ft = {
            sx: sub["employed_full_time"].sum() for sx, sub in g.groupby("sex")
        }
        pt = {
            sx: sub["employed_part_time"].sum() for sx, sub in g.groupby("sex")
        }
        for sx in ("males", "females"):
            states.append(
                {
                    "year": y,
                    "month": m,
                    "geography": geo,
                    "sex": sx,
                    "status_in_employment": status,
                    "employed_full_time": ft.get(sx, 0.0),
                    "employed_part_time": pt.get(sx, 0.0),
                }
            )
        states.append(
            {
                "year": y,
                "month": m,
                "geography": geo,
                "sex": "persons",
                "status_in_employment": status,
                "employed_full_time": sum(ft.values()),
                "employed_part_time": sum(pt.values()),
            }
        )
    st = pd.DataFrame(states)
    st["employed_total"] = st["employed_full_time"] + st["employed_part_time"]

    nat_wide = national.pivot_table(
        index=keys, columns="measure", values="value", aggfunc="first"
    ).reset_index()
    nat_wide.columns.name = None
    out = pd.concat([nat_wide, st], ignore_index=True)
    for c in arch_cols:
        if c not in out.columns:
            out[c] = pd.NA
    return out[arch_cols]


def _read_sem1(path: Path) -> pd.DataFrame:
    """Read SEM1's 'Data 1' pivot sheet -> long by state/sex/status/hours band.

    Returns:
        Columns: year, month, geography, sex, status_in_employment,
        employed_full_time, employed_part_time (absolute persons).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[next(s for s in wb.sheetnames if s.lower().startswith("data"))]
    it = ws.iter_rows(values_only=True)
    header = None
    for r in it:
        nn = [c for c in r if c is not None]
        if len(nn) >= 5 and any(
            isinstance(c, str) and str(c).strip() == "Month"  # pyrefly: ignore [unnecessary-type-conversion]
            for c in r  # pyrefly: ignore [unnecessary-type-conversion]
        ):
            header = [str(c).strip() if c is not None else "" for c in r]
            break
    # pyrefly: ignore [bad-argument-type]
    idx = {h: i for i, h in enumerate(header)}
    ci_month = idx["Month"]
    ci_sex = idx["Sex"]
    ci_state = next(
        i for h, i in idx.items() if h.startswith("State and territory")
    )
    ci_status = idx["Status in employment of main job"]
    ci_ft = next(
        i for h, i in idx.items() if h.startswith("Employed full-time")
    )
    ci_pt = next(
        i for h, i in idx.items() if h.startswith("Employed part-time")
    )
    recs = []
    for r in it:
        if not r or ci_month >= len(r) or r[ci_month] is None:
            continue
        # pyrefly: ignore [bad-argument-type]
        y, m = _year_month(pd.Timestamp(r[ci_month]).strftime("%Y-%m"))
        status = STATUS.get(str(r[ci_status]).strip().lower())
        sex = SEX.get(str(r[ci_sex]).strip().lower())
        if status is None or sex is None:
            continue
        ft = (
            # pyrefly: ignore [no-matching-overload]
            pd.to_numeric(r[ci_ft], errors="coerce")
            if ci_ft < len(r)
            else None
        )
        pt = (
            # pyrefly: ignore [no-matching-overload]
            pd.to_numeric(r[ci_pt], errors="coerce")
            if ci_pt < len(r)
            else None
        )
        recs.append(
            {
                "year": y,
                "month": m,
                "geography": str(r[ci_state]).strip(),
                "sex": sex,
                "status_in_employment": status,
                "employed_full_time": (ft or 0.0) * 1000.0,
                "employed_part_time": (pt or 0.0) * 1000.0,
            }
        )
    wb.close()
    return pd.DataFrame(recs)


# ── write ─────────────────────────────────────────────────────────────────────
def write_partitioned(df: pd.DataFrame, table: str, output_dir: Path) -> Path:
    """Write a table as all-STRING Snappy Parquet, hive-partitioned by year.

    Staging is all-STRING by Data Basis convention (dump_header stringifies the
    header BigQuery infers the staging schema from). Values pass through the
    architecture's real types first — so year serialises as "1978" not "1978.0"
    — then cast to string via arrow (never astype(str), which renders NULL as
    "nan" and defeats the dbt safe_cast). See [[project_dump_header_parquet_bug]].
    """
    arch = read_arch(table)
    order = [a["name"] for a in arch]
    typed = pa.schema(
        [pa.field(a["name"], PA[a["bigquery_type"]]) for a in arch]
    )
    string_schema = pa.schema([pa.field(a["name"], pa.string()) for a in arch])
    out = df[order].copy()
    tdir = output_dir / table
    total = 0
    for year, g in out.groupby("year", sort=True):
        # pyrefly: ignore [bad-argument-type]
        pdir = tdir / f"year={int(year)}"
        pdir.mkdir(parents=True, exist_ok=True)
        at = pa.Table.from_pandas(g, schema=typed, preserve_index=False)
        at = at.cast(string_schema)
        pq.write_table(at, pdir / "data.parquet", compression="snappy")
        total += len(g)
    log.info(f"{table}: {total:,} rows -> {tdir}")
    return tdir


def clean_all(input_dir: Path) -> dict[str, pd.DataFrame]:
    """Build all four Tier-1 tables from downloaded sources in ``input_dir``."""
    return {
        "labour_force_status": clean_labour_force_status(input_dir),
        "underutilisation": clean_underutilisation(input_dir),
        "hours_worked": clean_hours_worked(input_dir),
        "status_in_employment": clean_status_in_employment(input_dir),
    }
