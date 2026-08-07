"""Pure functions for the au_abs_cpi dataset: download + cleaning transform.

No Prefect imports here. The one-shot onboarding bootstrap
(models/au_abs_cpi/code/clean_data.py) and the recurring Prefect pipeline both
import these functions, so the cleaning transform lives in exactly one place.
"""

from __future__ import annotations

import functools
import os
from pathlib import Path

# pyrefly: ignore [untyped-import]
import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

from pipelines.datasets.au_abs_cpi.constants import constants

MEASURE_MAP = constants.MEASURE_MAP.value
COLUMNS = constants.COLUMNS.value
PERIOD_COL = constants.PERIOD_COL.value
YOY_LAG = constants.YOY_LAG.value

_ARROW_TYPE = {
    "year": pa.int64(),
    "quarter": pa.int64(),
    "month": pa.int64(),
    "region": pa.string(),
    "index_code": pa.string(),
    "serie_id": pa.string(),
    "index_name": pa.string(),
    "index_number": pa.float64(),
    "percentage_change_period": pa.float64(),
    "percentage_change_year": pa.float64(),
}


@functools.lru_cache(maxsize=1)
def _index_codes() -> dict[str, str]:
    """index_name -> ABS CL_CPI_INDEX code, from the bundled snapshot.

    For the 8 names that ABS repeats at two hierarchy levels with identical
    values (a group and its single child), the snapshot holds the higher-level
    (parent) code, chosen deterministically when it was generated.
    """
    path = Path(__file__).with_name(constants.INDEX_CODES_FILE.value)
    m = pd.read_csv(path, dtype=str)
    return dict(zip(m["index_name"], m["index_code"], strict=True))


# --------------------------------------------------------------------------- #
# Download
# --------------------------------------------------------------------------- #
def download_table(slug: str, file: str, out_dir: str, session=None) -> str:
    """Download one ABS xlsx to out_dir/<file>.xlsx and return the path."""
    import requests

    url = constants.BASE_URL.value.format(slug=slug, file=file)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{file}.xlsx")
    getter = session or requests
    resp = getter.get(url, headers=constants.HEADERS.value, timeout=120)
    resp.raise_for_status()
    with open(path, "wb") as fh:
        fh.write(resp.content)
    return path


def resolve_release_slug(session=None) -> str:
    """Resolve the current ABS CPI release slug (e.g. ``jun-2026``).

    The xlsx download URLs are dated by reference period; the ``latest-release``
    landing page carries links to the current period, so we read the slug from
    there rather than guessing it from the calendar (the release lags, and the
    monthly/quarterly cadences differ).
    """
    import re

    import requests

    getter = session or requests
    resp = getter.get(
        constants.LANDING_URL.value,
        headers=constants.HEADERS.value,
        timeout=120,
    )
    resp.raise_for_status()
    m = re.search(
        r"consumer-price-index-australia/([a-z]{3}-\d{4})/", resp.text
    )
    if not m:
        raise RuntimeError(
            "could not resolve ABS CPI release slug from landing page"
        )
    return m.group(1)


def download_all(out_dir: str, session=None) -> str:
    """Download every source workbook for the current release into out_dir."""
    slug = resolve_release_slug(session)
    files = sorted(
        {f for fs in constants.SOURCE_TABLES.value.values() for f in fs}
    )
    for f in files:
        download_table(slug, f, out_dir, session)
    return out_dir


# --------------------------------------------------------------------------- #
# Parse one ABS time-series workbook
# --------------------------------------------------------------------------- #
def _parse_desc(desc: str):
    """ "<measure> ; <item> ; <region> ;" -> (measure, item, region)."""
    # pyrefly: ignore [unnecessary-type-conversion]
    parts = [p.strip() for p in str(desc).split(";")]
    parts = [p for p in parts if p != ""]
    measure = parts[0] if len(parts) > 0 else None
    item = parts[1] if len(parts) > 1 else None
    region = parts[2] if len(parts) > 2 else None
    return measure, item, region


def _safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "..", "...", "na", "n.a.", "np"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_workbook(path: str) -> pd.DataFrame:
    """Return long records {measure_col, item, region, serie_id, year, month, value}
    for the kept measures of one ABS workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Series ID -> (measure_col, item, region) from the Index sheet.
    # Columns are located by header name (leading columns are merged/empty, so
    # "Series ID" is not at a fixed index).
    meta: dict[str, tuple] = {}
    ws = wb["Index"]
    sid_idx = None
    for row in ws.iter_rows(values_only=True):
        c0 = row[0]
        if c0 == "Data Item Description":
            sid_idx = row.index("Series ID")
            continue
        if sid_idx is None or c0 is None or str(c0).startswith("©"):
            continue
        # pyrefly: ignore [bad-argument-type]
        measure, item, region = _parse_desc(c0)
        # pyrefly: ignore [bad-argument-type]
        measure_col = MEASURE_MAP.get(measure)
        if measure_col is None:
            continue
        serie_id = row[sid_idx]
        # pyrefly: ignore [unsupported-operation]
        meta[serie_id] = (measure_col, item, region)

    records = []
    for sheet in (s for s in wb.sheetnames if s.lower().startswith("data")):
        ws = wb[sheet]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                if row[0] == "Series ID":
                    header = row  # column j -> series id
                continue
            date = row[0]
            if date is None or not hasattr(date, "year"):
                continue
            for j in range(1, len(header)):
                sid = header[j]
                if sid not in meta:
                    continue
                val = _safe_float(row[j])
                if val is None:
                    continue
                measure_col, item, region = meta[sid]
                records.append(
                    {
                        "measure_col": measure_col,
                        "item": item,
                        "region": region,
                        "serie_id": sid,
                        "year": int(date.year),
                        # pyrefly: ignore [missing-attribute]
                        "month": int(date.month),
                        "value": val,
                    }
                )
    wb.close()
    return pd.DataFrame.from_records(records)


# --------------------------------------------------------------------------- #
# Build one output frequency from its source workbooks
# --------------------------------------------------------------------------- #
def clean_frequency(frequency: str, input_dir: str) -> pd.DataFrame:
    files = constants.SOURCE_TABLES.value[frequency]
    period_col = PERIOD_COL[frequency]
    lag = YOY_LAG[frequency]

    raw = pd.concat(
        [parse_workbook(os.path.join(input_dir, f"{f}.xlsx")) for f in files],
        ignore_index=True,
    )
    # ABS labels each period by its final month: quarter = month // 3; month = month.
    raw["period"] = (
        raw["month"] // 3 if frequency == "quarterly" else raw["month"]
    )

    # The same (measure, region, item, year, period) can appear in >1 source table
    # (all-groups is repeated across T1/T3/T10 etc.). Values are identical; keep one.
    raw = raw.drop_duplicates(
        subset=["measure_col", "region", "item", "year", "period"],
        keep="first",
    )

    # Long -> wide by measure, keyed on (region, item, year, period).
    key = ["region", "item", "year", "period"]
    wide = raw.pivot_table(
        index=key, columns="measure_col", values="value", aggfunc="first"
    ).reset_index()
    for col in (
        "index_number",
        "percentage_change_period",
        "percentage_change_year",
    ):
        if col not in wide.columns:
            wide[col] = pd.NA

    # serie_id comes from the index-number series specifically.
    idx_ids = raw[raw["measure_col"] == "index_number"][
        [*key, "serie_id"]
    ].drop_duplicates(subset=key)
    wide = wide.merge(idx_ids, on=key, how="left")

    # Drop rows with no index level at all (a change series with no index is unusable).
    wide = wide[wide["index_number"].notna()].copy()

    # Fill missing percentage changes by computing from the index within each series.
    wide = wide.sort_values(["region", "item", "year", "period"]).reset_index(
        drop=True
    )
    grp = wide.groupby(["region", "item"], sort=False)["index_number"]
    computed_period = (wide["index_number"] / grp.shift(1) - 1.0) * 100.0
    computed_year = (wide["index_number"] / grp.shift(lag) - 1.0) * 100.0
    # Only fill where the shift lands on the immediately-preceding period / same
    # period one year earlier (guards against series with gaps).
    prev_year = wide.groupby(["region", "item"], sort=False)["year"].shift(1)
    prev_period = wide.groupby(["region", "item"], sort=False)["period"].shift(
        1
    )
    consecutive = _is_prev_period(frequency, wide, prev_year, prev_period)
    yoy_year = wide.groupby(["region", "item"], sort=False)["year"].shift(lag)
    yoy_period = wide.groupby(["region", "item"], sort=False)["period"].shift(
        lag
    )
    yoy_ok = (yoy_year == wide["year"] - 1) & (yoy_period == wide["period"])

    wide["percentage_change_period"] = wide["percentage_change_period"].where(
        wide["percentage_change_period"].notna(),
        computed_period.where(consecutive),
    )
    wide["percentage_change_year"] = wide["percentage_change_year"].where(
        wide["percentage_change_year"].notna(), computed_year.where(yoy_ok)
    )

    wide = wide.rename(columns={"item": "index_name", "period": period_col})

    # Attach the region-independent ABS item code (identifier of the product OL).
    wide["index_code"] = wide["index_name"].map(_index_codes())
    missing = sorted(
        wide.loc[wide["index_code"].isna(), "index_name"].unique()
    )
    if missing:
        raise ValueError(f"index_name(s) with no index_code: {missing}")

    out = wide[COLUMNS[frequency]].copy()
    out["year"] = out["year"].astype("int64")
    out[period_col] = out[period_col].astype("int64")
    out["index_number"] = pd.to_numeric(
        out["index_number"], errors="coerce"
    ).astype("float64")
    # ABS reports CPI movements to one decimal; round derived/published changes to
    # match that convention and avoid false precision (the index is 1-2 decimals).
    for c in ("percentage_change_period", "percentage_change_year"):
        out[c] = (
            pd.to_numeric(out[c], errors="coerce").astype("float64").round(1)
        )
    for c in ("region", "index_code", "serie_id", "index_name"):
        out[c] = out[c].astype("string").astype("object")
    return out.sort_values(
        ["region", "index_name", "year", period_col]
    ).reset_index(drop=True)


def _is_prev_period(frequency, df, prev_year, prev_period):
    """True where the previous row is the immediately preceding period.

    Called before the generic ``period`` column is renamed to quarter/month.
    """
    last = 12 if frequency == "monthly" else 4
    same_year = (prev_year == df["year"]) & (prev_period == df["period"] - 1)
    year_wrap = (
        (prev_year == df["year"] - 1)
        & (df["period"] == 1)
        & (prev_period == last)
    )
    return same_year | year_wrap


def clean_all(input_dir: str, output_dir: str) -> dict[str, str]:
    """Build every table into partitioned parquet under output_dir.

    Returns a mapping of table slug to its partition root, plus
    ``"max_year_month"`` — the latest ``"YYYY-MM"`` in the monthly table, which
    drives the source-update poll.
    """
    result: dict[str, str] = {}
    max_ym = None
    for freq in constants.SOURCE_TABLES.value:
        # pyrefly: ignore [unnecessary-type-conversion]
        df = clean_frequency(freq, str(input_dir))
        # pyrefly: ignore [unnecessary-type-conversion]
        write_partitioned(df, freq, str(output_dir))
        result[freq] = str(Path(output_dir) / freq)
        if freq == "monthly":
            last = df.sort_values(["year", "month"]).iloc[-1]
            max_ym = f"{int(last['year']):04d}-{int(last['month']):02d}"
    # pyrefly: ignore [unsupported-operation]
    result["max_year_month"] = max_ym
    return result


# --------------------------------------------------------------------------- #
# Write partitioned parquet (all-STRING, hive-partitioned by year)
# --------------------------------------------------------------------------- #
def write_partitioned(df: pd.DataFrame, table: str, out_dir: str) -> int:
    """Write a table as all-STRING Snappy parquet, hive-partitioned by year.

    Staging is all-STRING by Data Basis convention — the dbt model ``safe_cast``s
    every column, and the pipeline's ``upload_to_gcs`` infers the staging schema
    from a stringified header (``dump_header``), which rejects typed parquet. So
    values pass through the architecture's real arrow types first (``year`` ->
    ``"1948"`` not ``"1948.0"``, NULL preserved) and are then cast to string via
    arrow — never ``astype(str)``, which would render NULL as ``"nan"`` and
    defeat ``safe_cast``. The partition column ``year`` is kept in the file, as
    in ``us_bls_cpi``.
    """
    cols = COLUMNS[table]
    typed_schema = pa.schema([(c, _ARROW_TYPE[c]) for c in cols])
    string_schema = pa.schema([(c, pa.string()) for c in cols])
    n = 0
    for year, part in df.groupby("year", sort=True):
        # pyrefly: ignore [bad-argument-type]
        dest = Path(out_dir) / table / f"year={int(year)}"
        dest.mkdir(parents=True, exist_ok=True)
        at = pa.Table.from_pandas(
            part[cols], schema=typed_schema, preserve_index=False
        )
        at = at.cast(string_schema)
        pq.write_table(at, dest / "data.parquet", compression="snappy")
        n += len(part)
    return n
